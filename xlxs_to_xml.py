
import openpyxl
import xml.dom.minidom


full_name = "./中国电信4G设备统一网管 - 4G性能指标0407.xlsx"

global_list = []

def get_objname_convert(obj : str):
    return 'Cell' if obj == '小区' else obj 

def get_stat_convert(stat : str) :
    if stat == '最新':
        return '5'
    if stat == '平均':
        return '0'
    if stat == '最大':
        return '1'
    if stat == '累积':
        return '2'
    return str(stat)

def get_out_convert(sur : str):
    integer_tuple = ('dBm', 'Mbps', 'MByte', '包', '次', '个', '毫秒', '秒', '瓦', '用户数/秒', 'MByte、个（包）', 'Mbyte、个（包）')
    real_tuple = ('bps/Hz', '百分比')
    if sur in integer_tuple:
        return '1'
    if sur in real_tuple:
        return '0'
    return '3'

def load_xls_data():
    try:
        excel_workbook = openpyxl.load_workbook(full_name)
        work_sheet = excel_workbook.worksheets[0]
        sheet_rows = work_sheet.max_row
        # 读取xls的内容， cell(row, col).value
        #暂时只用到obj, stat,sur
        for i in range(4, sheet_rows):
            local_temp_dict = {'kpi'        : str(work_sheet.cell(i, 7).value),
                            'kpi_ch'     : str(work_sheet.cell(i, 5).value), 
                            'kpi_sub_ch' : str(work_sheet.cell(i, 6).value), 
                            'source'     : str(work_sheet.cell(i, 9).value),
                            'obj'        : get_objname_convert(str(work_sheet.cell(i, 10).value )), 
                            'stat'       : get_stat_convert(str(work_sheet.cell(i, 11).value)), 
                            'sur'        : get_out_convert(str(work_sheet.cell(i, 12).value)) }
            global_list.append(local_temp_dict)
    except Exception as result:
        raise Exception('-%s- <%s>'%(str(result.__traceback__.tb_lineno),result))
    
def write_xml():
    try:
        impl = xml.dom.minidom.getDOMImplementation()
        
        #设置根结点configuration
        dom = impl.createDocument(None, 'configuration', None)
        root = dom.documentElement
        #添加global空节点
        node_global = dom.createElement('global')
        root.appendChild(node_global)
        
        #dom.createElement(node_name) 创建节点 , 名称 node_name
        #node.setAttribute('key', 'value') 创建节点的属性
        #createTextNode(value) 创建节点的内容
        #node.appendChild(sub_node) 添加子节点
        for i in range(len(global_list)):
            pmMeas = dom.createElement('pmMeas')
            pmMeas.setAttribute('measName', global_list[i]['kpi'])
            pmMeas.setAttribute('measObjName', global_list[i]['obj'])
            pmMeas.setAttribute('statType', global_list[i]['stat'])
            pmMeas.setAttribute('outType', global_list[i]['sur'])
            root.appendChild(pmMeas)
            
            countId = dom.createElement('countId')
            countId.setAttribute('counterType', '0')
            countIdText = dom.createTextNode('0')
            countId.appendChild(countIdText)
            pmMeas.appendChild(countId)
            
        
        
        with open('emplist.xml', 'w') as file_object:
            dom.writexml(file_object, addindent='\t', newl='\n')
            file_object.write('\n')
    except Exception as result:
        raise Exception('-%s- <%s>'%(str(result.__traceback__.tb_lineno),result))    


if __name__ == "__main__":
    print ("hello world")
    load_xls_data()
    write_xml()
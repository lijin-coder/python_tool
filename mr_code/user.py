
import glob
import os
import xml
import re
from lxml import etree
import time
import math
import openpyxl
import mr_globel as gl
import mr_utils
from mr_qt import write_info


def test51_file_integrity():
    mr_utils.test_out_data_item_header("test_51")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    try:
        #统计预期的文件总数,计算: 测量总时长/测量周期
        predict_file_num =  int(gl.TEST_CONF['test_total_time'])*3600 / (60 * int(gl.MR_CONF['UploadPeriod']) )

        mr_integrity_flag = 0

        #MRS文件数量统计
        file_list_mro = glob.glob(gl.MR_TEST_PATH + '*MRO*.xml')
        for entity in file_list_mro:
            if os.path.getsize(entity) == 0 or mr_utils.MR_xml_file_name_accuracy(entity) == False:
                mr_integrity_flag = mr_integrity_flag | (0x1 << 1)
                break

        file_list_mre = glob.glob(gl.MR_TEST_PATH + '*MRE*.xml')
        for entity in file_list_mre:
            if os.path.getsize(entity) == 0 or mr_utils.MR_xml_file_name_accuracy(entity) == False:
                mr_integrity_flag = mr_integrity_flag | (0x1 << 2)
                break

        file_list_mrs = glob.glob(gl.MR_TEST_PATH + '*MRS*.xml')
        for entity in file_list_mrs:
            if os.path.getsize(entity) == 0 or mr_utils.MR_xml_file_name_accuracy(entity) == False:
                mr_integrity_flag = mr_integrity_flag | (0x1 << 3)
                break




        with open(gl.OUT_PATH, 'a') as file_object:
            file_object.write(date_time + " | ")
            file_object.write(gl.TEST_CONF['test_total_time'] + " | ")
            file_object.write(str(len(gl.TEST_CONF['enbid'].split(','))) + " | ")
            file_object.write(str(predict_file_num * (len(gl.MR_CONF['MeasureType'].split(',')))) + ' | ')
            file_object.write(str(len(file_list_mrs)) + ' | ')
            file_object.write(str(len(file_list_mro)) + ' | ')
            file_object.write(str(len(file_list_mre)) + ' | ')
            file_object.write(str(mr_integrity_flag & (0x1 << 3) == 0 and \
                                  (predict_file_num == len(file_list_mrs) and  predict_file_num == len(gl.MR_DICT)) \
                                  if re.search(r'MRS',gl.MR_CONF['MeasureType']) != None \
                                  else  (len(file_list_mrs) == 0) ) + ' | ')
            file_object.write(str(mr_integrity_flag & (0x1 << 1) == 0 and \
                                  (predict_file_num == len(file_list_mro) and predict_file_num == len(gl.MR_DICT) ) \
                                  if re.search('MRO', gl.MR_CONF['MeasureType']) != None \
                                  else (len(file_list_mro == 0))) + ' | ')
            file_object.write(str(mr_integrity_flag & (0x1 << 2) == 0 and \
                                  (predict_file_num == len(file_list_mre) and predict_file_num == len(gl.MR_DICT))\
                                  if re.search('MRE', gl.MR_CONF['MeasureType']) != None \
                                  else (len(file_list_mre == 0)) ) + ' | ')
    except Exception as result:
        raise Exception('<%s>'%(result))


def test52_file_integrity():
    mr_utils.test_out_data_item_header("test_52")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    temp_flag_dict = {"MR.RSRP":0, "MR.RSRQ":0, "MR.SinrUL":0, "MR.PowerHeadRoom":0, "filename":[]}
    temp_flag_integrity = 0
    out_dict = {}
    output_temp_dict = {}
    out_sort_list = ["MR.RSRP", "MR.RSRQ","MR.PowerHeadRoom", "MR.SinrUL" ]

    file_list_mrs = glob.glob(gl.MR_TEST_PATH + '*MRS*.xml')

    for mrs_file_entity in file_list_mrs:

        try:
            mrs_temp_dom = xml.dom.minidom.parse(mrs_file_entity)
            mrs_root = mrs_temp_dom.documentElement
            if out_dict.__contains__(mrs_file_entity) == False:
                out_dict[mrs_file_entity] = {"MR.RSRP": {'num':0, 'list':[]}, "MR.RSRQ":{'num':0, 'list':[]}, "MR.SinrUL":{'num':0, 'list':[]}, "MR.PowerHeadRoom":{'num':0, 'list':[]}}
            for enb_entity in mrs_root.getElementsByTagName('eNB'):
                enbid = int(enb_entity.getAttribute('id'))

                for measurement_entity in enb_entity.getElementsByTagName('measurement'):
                    mrName = measurement_entity.getAttribute('mrName')
                    object_list = measurement_entity.getElementsByTagName('object')
                    for mr_name_entity in out_dict[mrs_file_entity]:
                        if mr_name_entity == mrName:
                            for object_entity in object_list:
                                eci_id = int(object_entity.getAttribute('id').split(':')[0])
                                cellid_ret_list = mr_utils.is_cell_id_exist(eci_id)
                                if  cellid_ret_list[0] == False:
                                    mr_utils.out_text_dict_append_list(output_temp_dict, mrs_file_entity, '[%s]-[cellid:%d not in list] '%(mrName, cellid_ret_list[1]))
                                else:
                                    if cellid_ret_list[1] not in out_dict[mrs_file_entity][mr_name_entity]['list']:
                                        out_dict[mrs_file_entity][mr_name_entity]['num'] += 1
                                        out_dict[mrs_file_entity][mr_name_entity]['list'].append(cellid_ret_list[1])
                            break
        except Exception as result:
            raise Exception('<%s> : %s '%(result, mrs_file_entity))


    with open(gl.OUT_PATH, 'a') as file_object:
        cell_num = len(gl.TEST_CONF['cellid'].split(','))
        file_object.write(date_time + " | ")
        file_object.write(gl.TEST_CONF['test_total_time'] + " | ")
        for file_name in out_dict:
            file_object.write('\n=====> ' + file_name + ': |')
            temp_flag_test = 0
            for mr_name in out_sort_list:
                cell_num_mrs_item = out_dict[file_name][mr_name]['num']
                file_object.write(str(cell_num) + " | ")
                if (gl.MR_CONF['MeasureItems'] == 'all' or re.search(mr_name, gl.MR_CONF['MeasureItems']) != None) and cell_num_mrs_item != cell_num:
                    temp_flag_test = 1
                    mr_utils.out_text_dict_append_list(output_temp_dict, mrs_file_entity, '[%s]=<cell num not match>:[%d(mrs)]!=[%d(target)]'%(mr_name, cell_num_mrs_item, cell_num))
                if mr_utils.is_mr_item_need_exist(mr_name) == False and cell_num_mrs_item != 0:
                    temp_flag_test = 1
                    mr_utils.out_text_dict_append_list(output_temp_dict, mrs_file_entity, '[{0}]=<MeasureItems not have {0}>:cellnum=[{1} ->target:0]'.format(mr_name, cell_num_mrs_item))
            file_object.write(str(temp_flag_test == 0))
            if output_temp_dict.__contains__(file_name) == True and len(output_temp_dict[file_name]) != 0:
                file_object.write('\n===> error:\n')
                for i in range(len(output_temp_dict[file_name])):
                    file_object.write('\t\t' + output_temp_dict[file_name][i] + '\n')


def test53_file_integrity():
    mr_utils.test_out_data_item_header("test_53")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    file_list_mrs = glob.glob(gl.MR_TEST_PATH + '*MRS*.xml')


    text_consistent = {}

    subfram_dict = {}

    for file_mrs_name in file_list_mrs:
        try:
            subfram_dict[file_mrs_name] = {}
            for cell_id in gl.TEST_CONF['cellid'].split(','):
                subfram_dict[file_mrs_name][cell_id] = {}
                for dict_name_key in gl.MR_CONF['SubFrameNum'].split(','):
                    subfram_dict[file_mrs_name][cell_id][dict_name_key] = 0
            subfram_dict[file_mrs_name][cell_id]['consistent'] = 0
            mrs_dom = xml.dom.minidom.parse(file_mrs_name)
            mrs_root = mrs_dom.documentElement
            measurement_list = mrs_root.getElementsByTagName('measurement')
            for measurement_entity in measurement_list:
                if measurement_entity.getAttribute('mrName') == 'MR.ReceivedIPower' :
                    object_list = measurement_entity.getElementsByTagName('object')
                    for object_entity in object_list:
                        is_consistent = 0
                        eci_id = int(object_entity.getAttribute('id').split(':')[0])
                        cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                        subfram_id = object_entity.getAttribute('id').split(':')[2]
                        test_count = 0
                        if subfram_dict[file_mrs_name].__contains__(str(cell_id_ret_list[1])) == False:
                            mr_utils.out_text_dict_append_list(text_consistent, file_mrs_name, '[MR.ReceivedIPower]:<cell_id[%d] not in list>'%(cell_id_ret_list[1]))
                        else:
                            for subfram_key in subfram_dict[file_mrs_name][str(cell_id_ret_list[1])]:
                                if subfram_key == subfram_id:
                                    subfram_dict[file_mrs_name][str(cell_id_ret_list[1])][subfram_key] += 1
                                    test_count = 1
                                    break
                            if test_count == 0:
                                mr_utils.out_text_dict_append_list(text_consistent, file_mrs_name, '[MR.ReceivedIPower]:<surplus rip [cellid:%d]-[%s]>'%(cell_id_ret_list[1],subfram_id))
                                subfram_dict[file_mrs_name][str(cell_id_ret_list[1])]['consistent'] = 1
                    break
            for cell_id_entity in subfram_dict[file_mrs_name]:
                for subfram_entity in gl.MR_CONF['SubFrameNum'].split(','):
                    if subfram_dict[file_mrs_name][cell_id_entity][subfram_entity] != 1 :
                        mr_utils.out_text_dict_append_list(text_consistent, file_mrs_name, '[MR.ReceivedIPower]:<confusion rip [cellid:%s]-[%s]>'%(cell_id_entity,subfram_entity))
                        subfram_dict[file_mrs_name][cell_id_entity]['consistent'] = 1
        except Exception as result:
            raise Exception('<%s> : [%s]'%(result, file_mrs_name))
    with open(gl.OUT_PATH, 'a') as file_object:
        file_object.write(date_time + " | ")
        file_object.write(gl.TEST_CONF['test_total_time'] + " | \n")
        for mrs_file in subfram_dict:
            file_object.write('======>%s:'%(mrs_file))
            for cell_id_entity in subfram_dict[mrs_file]:
                file_object.write(' | cellid:[%s]-ripnum:[%d] | %s | \n'%(cell_id_entity, len(subfram_dict[mrs_file][cell_id_entity]) - 1,
                        str(subfram_dict[mrs_file][cell_id_entity]['consistent'] == 0)))
            if text_consistent.__contains__(mrs_file) == True and len(text_consistent[mrs_file]) != 0:
                file_object.write('====>error:\n')
                for i in range(len(text_consistent[mrs_file])):
                    file_object.write(text_consistent[mrs_file][i] + '\n')


def test54_file_integrity():
    mr_utils.test_out_data_item_header("test_54")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    file_list_mrs = glob.glob(gl.MR_TEST_PATH + '*MRS*.xml')
    prbnum_list = []
    if re.search(r',', gl.MR_CONF['PrbNum']) == None:
        for i in range(int(gl.MR_CONF['PrbNum'].split('....')[0]), 1 + int(gl.MR_CONF['PrbNum'].split('....')[1])):
            prbnum_list.append(str(i))
    else:
        prbnum_list = gl.MR_CONF['PrbNum'].split(',')

    ripprb_num = len(gl.MR_CONF['SubFrameNum'].split(',')) * len(prbnum_list)
    is_consistence = 0
    text_consistence = {}
    mrs_ripprb_dict = {}

    for file_name in file_list_mrs:
        try:
            mrs_ripprb_dict[file_name] = {}
            mrs_dom = xml.dom.minidom.parse(file_name)
            mrs_root = mrs_dom.documentElement
            measurement_list = mrs_root.getElementsByTagName('measurement')
            for measurement_entity in measurement_list:
                if measurement_entity.getAttribute('mrName') == 'MR.RIPPRB':
                    object_list = measurement_entity.getElementsByTagName('object')
                    for object_entity in object_list:
                        id_str_list = object_entity.getAttribute('id').split(':')
                        eci_id = int(id_str_list[0])
                        rip = id_str_list[2]
                        prb = id_str_list[3]
                        cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)

                        if mrs_ripprb_dict[file_name].__contains__(str(cell_id_ret_list[1])) == False:
                            mrs_ripprb_dict[file_name][str(cell_id_ret_list[1])] = []
                        if re.search(rip, gl.MR_CONF['SubFrameNum']) != None and prb in prbnum_list:
                            mrs_ripprb_dict[file_name][str(cell_id_ret_list[1])].append(rip + ':' + prb)
                        else:
                            mr_utils.out_text_dict_append_list(text_consistence, file_name, '<ripprb not match>:[%s:%s]'%(rip,prb))
                    break
        except Exception as result:
            raise Exception('<%s> : [%s]'%(result, file_name))
    with open(gl.OUT_PATH, 'a') as file_object:
        file_object.write(date_time + " | ")
        file_object.write(gl.TEST_CONF['test_total_time'] + " | \n")
        for file_name in mrs_ripprb_dict:
            file_object.write('====>%s: \n\t\t'%(file_name))
            for cell_id_str in mrs_ripprb_dict[file_name]:
                ripnum = len(mrs_ripprb_dict[file_name][cell_id_str])
                file_object.write('<[cell:%s]:[%d] | %s>\n'%(cell_id_str, ripnum, ripnum == ripprb_num ))
            if text_consistence.__contains__(file_name) == True:
                file_object.write('error:\n')
                for i in range(len(text_consistence[file_name])):
                    file_object.write(text_consistence[file_name][i] + '\n')


def test55_file_integrity():
    mr_utils.test_out_data_item_header("test_55")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    out_list_mro = ['MR.LteScRSRP', 'MR.LteScRSRQ', 'MR.LteScPHR']
    out_list_mrs = ['MR.RSRP', 'MR.RSRQ', 'MR.PowerHeadRoom']
    for time_list_entity in gl.MR_DICT:
        mro_count_dict = {'MR.LteScRSRP':{}, 'MR.LteScRSRQ':{}, 'MR.LteScPHR':{}}
        mrs_count_dict = {'MR.RSRP':{}, 'MR.RSRQ':{}, 'MR.PowerHeadRoom':{}}
        pos_dict = {'MR.LteScRSRP': {'pos':2}, 'MR.LteScRSRQ': {'pos':3}, 'MR.LteScPHR':{'pos':4}}
        try:
            mro_measurement_list = []
            mrs_measurement_list = []
            if re.search(r'MRO', gl.MR_CONF['MeasureType']) != None:
                mro_file_name = gl.MR_DICT[time_list_entity][0]['MRO']
                mro_dom = xml.dom.minidom.parse(mro_file_name)
                mro_root = mro_dom.documentElement
                mro_measurement_list = mro_root.getElementsByTagName('measurement')

            if re.search(r'MRS', gl.MR_CONF['MeasureType']) != None:
                mrs_file_name = gl.MR_DICT[time_list_entity][0]['MRS']
                mrs_dom = xml.dom.minidom.parse(mrs_file_name)
                mrs_root = mrs_dom.documentElement
                mrs_measurement_list = mrs_root.getElementsByTagName('measurement')

            temp_pos = 2
            for mro_measurement_entity in mro_measurement_list:
                smr_list = mro_measurement_entity.getElementsByTagName('smr')
                for smr_entity in smr_list:
                    smr_str = smr_entity.firstChild.data
                    mr_utils.get_mr_item_pos(pos_dict, smr_str)
                    for mro_mr_Name in mro_count_dict:
                        for smr_name_entity in smr_str.split(' '):
                            item_count = 0
                            if mro_mr_Name == smr_name_entity:
                                for object_entity in mro_measurement_entity.getElementsByTagName('object'):
                                    eci_id = int(object_entity.getAttribute('id'))
                                    cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                                    object_id = str(cell_id_ret_list[1])
                                    if mro_count_dict[mro_mr_Name].__contains__(object_id) == False:
                                        mro_count_dict[mro_mr_Name][object_id] = 0
                                    value_list = object_entity.getElementsByTagName('v')
                                    for value_entity in value_list:
                                        if value_entity.firstChild.data.split(' ')[pos_dict[mro_mr_Name]['pos']].isdigit() == True :
                                            mro_count_dict[mro_mr_Name][object_id] += 1
                                    item_count += 1
                                    #if mro_count_dict[mro_mr_Name][object_id] != item_count:
                                    #print (mro_mr_Name + " | " + str(mro_count_dict[mro_mr_Name][object_id]))
                                break
                        temp_pos += 1

            for mrs_measurement_entity in mrs_measurement_list:
                for mrs_mr_Name in mrs_count_dict:
                    if mrs_measurement_entity.getAttribute('mrName') == mrs_mr_Name:
                        for object_entity in mrs_measurement_entity.getElementsByTagName('object'):
                            eci_id = int(object_entity.getAttribute('id'))
                            cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                            object_id = str(cell_id_ret_list[1])
                            temp_count = 0
                            for value_entity in object_entity.getElementsByTagName('v'):
                                temp_count += mr_utils.add_digital_string(value_entity.firstChild.data)

                            mrs_count_dict[mrs_mr_Name][object_id] = temp_count

            with open(gl.OUT_PATH, 'a') as file_object:
                file_object.write(time_list_entity + ' : \n')
                file_object.write(date_time + " | ")
                file_object.write(gl.TEST_CONF['test_total_time'] + " | ")

                file_object.write('[MRO文件上报采样点数]' + " | ")
                for mr_name in out_list_mro:
                    file_object.write('{')
                    for object_id in mro_count_dict[mr_name]:
                        file_object.write(' [%s]=%s '%(object_id, str(mro_count_dict[mr_name][object_id])) )
                    file_object.write('} | ')
                file_object.write('\n' + " " * (len(date_time) + len(gl.TEST_CONF['test_total_time']) + 6))

                file_object.write('[MRS文件统计采样点数]' + " | ")
                for mr_name in out_list_mrs:
                    file_object.write('{')
                    for object_id in mrs_count_dict[mr_name]:
                        file_object.write(' [%s]=%s '%(object_id, str(mrs_count_dict[mr_name][object_id])))
                    file_object.write('} | ')
                file_object.write('\n' + " " * (len(date_time) + len(gl.TEST_CONF['test_total_time']) + 6))

                file_object.write('[统计采样点数是否完整]' + " | ")
                file_object.write(str(mro_count_dict['MR.LteScRSRP'] == mrs_count_dict['MR.RSRP']) + ' | ')
                file_object.write(str(mro_count_dict['MR.LteScRSRQ'] == mrs_count_dict['MR.RSRQ']) + ' | ')
                file_object.write(str(mro_count_dict['MR.LteScPHR'] == mrs_count_dict['MR.PowerHeadRoom']) + ' | \n')
        except Exception as result:
            raise Exception('<%s> MRO:[%s] MRS:[%s]'%(result, gl.MR_DICT[time_list_entity][0]['MRO'],gl.MR_DICT[time_list_entity][0]['MRS']))


def test56_file_integrity():
    mr_utils.test_out_data_item_header("test_56")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    out_mro_list = ['MR.LteScRIP', 'MR.LteScSinrUL']
    out_mrs_list = ['MR.ReceivedIPower', 'MR.SinrUL']
    for time_list_entity in gl.MR_DICT:
        try:
            mro_count_dict = {'MR.LteScRIP':{}, 'MR.LteScSinrUL':{}}
            mrs_count_dict = {'MR.ReceivedIPower':{}, 'MR.SinrUL':{}}
            mro_measurement_list = []
            mrs_measurement_list = []
            if re.search(r'MRO', gl.MR_CONF['MeasureType']) != None:
                mro_file_name = gl.MR_DICT[time_list_entity][0]['MRO']
                mro_dom = xml.dom.minidom.parse(mro_file_name)
                mro_root = mro_dom.documentElement
                mro_measurement_list = mro_root.getElementsByTagName('measurement')
            if re.search(r'MRS', gl.MR_CONF['MeasureType']) != None:
                mrs_file_name = gl.MR_DICT[time_list_entity][0]['MRS']
                mrs_dom = xml.dom.minidom.parse(mrs_file_name)
                mrs_root = mrs_dom.documentElement
                mrs_measurement_list = mrs_root.getElementsByTagName('measurement')

            for mr_name in out_mro_list:
                for mro_measurement_entity in mro_measurement_list:
                    for smr_entity in mro_measurement_entity.getElementsByTagName('smr'):
                        if mr_name in smr_entity.firstChild.data.split(' ') :
                            for object_entity in mro_measurement_entity.getElementsByTagName('object'):
                                object_value = object_entity.getAttribute('id')
                                eci_id = int(object_value.split(':')[0])
                                cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                                object_id = str(cell_id_ret_list[1])
                                object_prb = ""
                                if mr_name == 'MR.LteScRIP':
                                    object_prb = object_value.split(':')[2]

                                if mro_count_dict[mr_name].__contains__(object_id) == False :
                                    if mr_name == 'MR.LteScRIP':
                                        mro_count_dict[mr_name][object_id] = {}
                                    else:
                                        mro_count_dict[mr_name][object_id] = 0


                                if mr_name == 'MR.LteScRIP' and mro_count_dict[mr_name][object_id].__contains__(object_prb) == False:
                                    mro_count_dict[mr_name][object_id][object_prb] = 0

                                for value_entity in object_entity.getElementsByTagName('v'):
                                    if mr_name == 'MR.LteScRIP' and value_entity.firstChild.data.isdigit() == True:
                                        mro_count_dict[mr_name][object_id][object_prb] += 1
                                    elif mr_name == 'MR.LteScSinrUL' and value_entity.firstChild.data.split(' ')[5].isdigit() == True:
                                            #TODO:在这里还需要商量一下,如果在MRO中没有统计,那么在MRS中,采样点数目是对不上的
                                        #if int(value_entity.firstChild.data.split(' ')[5]) != 0 :
                                        mro_count_dict[mr_name][object_id] += 1

            for mr_name in out_mrs_list:
                for measurement_entity in mrs_measurement_list:
                    if mr_name == measurement_entity.getAttribute('mrName'):
                        for object_entity in measurement_entity.getElementsByTagName('object'):
                            eci_id = int(object_entity.getAttribute('id').split(':')[0])
                            cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                            object_id = str(cell_id_ret_list[1])
                            object_prb = ''
                            if mr_name == 'MR.ReceivedIPower':
                                object_prb = object_entity.getAttribute('id').split(':')[2]

                            temp_num = 0
                            for value_entity in object_entity.getElementsByTagName('v'):
                                temp_num += mr_utils.add_digital_string(value_entity.firstChild.data)

                            if mr_name == 'MR.ReceivedIPower':
                                if mrs_count_dict[mr_name].__contains__(object_id) == False:
                                        mrs_count_dict[mr_name][object_id] = {}
                                if mrs_count_dict[mr_name][object_id].__contains__(object_prb) == False:
                                    mrs_count_dict[mr_name][object_id][object_prb] = 0
                                mrs_count_dict[mr_name][object_id][object_prb] = temp_num
                            elif mr_name == 'MR.SinrUL':
                                if mrs_count_dict[mr_name].__contains__(object_id) == False:
                                    mrs_count_dict[mr_name][object_id] = 0
                                mrs_count_dict[mr_name][object_id] = temp_num

            with open(gl.OUT_PATH, 'a') as file_object:
                file_object.write(time_list_entity + ' : \n')
                file_object.write(date_time + " | ")
                file_object.write(gl.TEST_CONF['test_total_time'] + " | ")

                file_object.write('[MRO文件上报采样点数]' + " | ")
                for mr_name in out_mro_list:
                    for object_id in mro_count_dict[mr_name]:
                        if mr_name == 'MR.LteScRIP':
                            file_object.write('[' + object_id + ']={' )
                            for object_prb in mro_count_dict[mr_name][object_id]:
                                file_object.write('[' + object_prb + ']=' + str(mro_count_dict[mr_name][object_id][object_prb]) + ' ')
                            file_object.write('} | ')
                        if mr_name == 'MR.LteScSinrUL':
                            file_object.write('[' + object_id + ']=' + str(mro_count_dict[mr_name][object_id]) + ' | ')

                file_object.write('\n' + " " * (len(date_time) + len(gl.TEST_CONF['test_total_time']) + 6))

                file_object.write('[MRS文件统计采样点数]' + " | ")
                for mr_name in out_mrs_list:
                    for object_id in mrs_count_dict[mr_name]:
                        if mr_name == 'MR.ReceivedIPower':
                            file_object.write('[' + object_id + ']={' )
                            for object_prb in mrs_count_dict[mr_name][object_id]:
                                file_object.write('[' + object_prb + ']=' + str(mrs_count_dict[mr_name][object_id][object_prb]) + ' ')
                            file_object.write('} | ')
                        if mr_name == 'MR.SinrUL':
                            file_object.write('[' + object_id + ']=' + str(mrs_count_dict[mr_name][object_id]) + ' | ')

                file_object.write('\n' + " " * (len(date_time) + len(gl.TEST_CONF['test_total_time']) + 6))

                test_flag = 0
                for object_id in mrs_count_dict['MR.ReceivedIPower']:
                    for object_prb in mrs_count_dict['MR.ReceivedIPower'][object_id]:
                        if mro_count_dict['MR.LteScRIP'][object_id].__contains__(object_prb) == True and mro_count_dict['MR.LteScRIP'][object_id][object_prb] != mrs_count_dict['MR.ReceivedIPower'][object_id][object_prb]:
                            test_flag = 1
                file_object.write('[统计采样点数是否完整]' + " | ")
                file_object.write(str(0 == test_flag) + ' | ')
                file_object.write(str(mro_count_dict['MR.LteScSinrUL'] == mrs_count_dict['MR.SinrUL']) + ' | \n')
        except Exception as result:
            raise Exception('<%s> : MRO:[%s] MRS:[%s]'%(result, gl.MR_DICT[time_list_entity][0]['MRO'], gl.MR_DICT[time_list_entity][0]['MRS']))

def test57_file_integrity():
    mr_utils.test_out_data_item_header("test_57")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    sheet_row_dict = {'MR.RSRP':[3,49], 'MR.RSRQ':[51,67],'MR.PowerHeadRoom':[114,176], 'MR.ReceivedIPower':[178,229],'MR.SinrUL':[807,842], 'MR.RIPPRB':[844, 895]}

    rsrp_smr = 'MR.RSRP.00 MR.RSRP.01 MR.RSRP.02 MR.RSRP.03 MR.RSRP.04 MR.RSRP.05 MR.RSRP.06 MR.RSRP.07 MR.RSRP.08 MR.RSRP.09 MR.RSRP.10 MR.RSRP.11 MR.RSRP.12 MR.RSRP.13 MR.RSRP.14 MR.RSRP.15 MR.RSRP.16 MR.RSRP.17 MR.RSRP.18 MR.RSRP.19 MR.RSRP.20 MR.RSRP.21 MR.RSRP.22 MR.RSRP.23 MR.RSRP.24 MR.RSRP.25 MR.RSRP.26 MR.RSRP.27 MR.RSRP.28 MR.RSRP.29 MR.RSRP.30 MR.RSRP.31 MR.RSRP.32 MR.RSRP.33 MR.RSRP.34 MR.RSRP.35 MR.RSRP.36 MR.RSRP.37 MR.RSRP.38 MR.RSRP.39 MR.RSRP.40 MR.RSRP.41 MR.RSRP.42 MR.RSRP.43 MR.RSRP.44 MR.RSRP.45 MR.RSRP.46 MR.RSRP.47 '
    rsrq_smr = 'MR.RSRQ.00 MR.RSRQ.01 MR.RSRQ.02 MR.RSRQ.03 MR.RSRQ.04 MR.RSRQ.05 MR.RSRQ.06 MR.RSRQ.07 MR.RSRQ.08 MR.RSRQ.09 MR.RSRQ.10 MR.RSRQ.11 MR.RSRQ.12 MR.RSRQ.13 MR.RSRQ.14 MR.RSRQ.15 MR.RSRQ.16 MR.RSRQ.17 '
    rip_smr = 'MR.ReceivedIPower.00 MR.ReceivedIPower.01 MR.ReceivedIPower.02 MR.ReceivedIPower.03 MR.ReceivedIPower.04 MR.ReceivedIPower.05 MR.ReceivedIPower.06 MR.ReceivedIPower.07 MR.ReceivedIPower.08 MR.ReceivedIPower.09 MR.ReceivedIPower.10 MR.ReceivedIPower.11 MR.ReceivedIPower.12 MR.ReceivedIPower.13 MR.ReceivedIPower.14 MR.ReceivedIPower.15 MR.ReceivedIPower.16 MR.ReceivedIPower.17 MR.ReceivedIPower.18 MR.ReceivedIPower.19 MR.ReceivedIPower.20 MR.ReceivedIPower.21 MR.ReceivedIPower.22 MR.ReceivedIPower.23 MR.ReceivedIPower.24 MR.ReceivedIPower.25 MR.ReceivedIPower.26 MR.ReceivedIPower.27 MR.ReceivedIPower.28 MR.ReceivedIPower.29 MR.ReceivedIPower.30 MR.ReceivedIPower.31 MR.ReceivedIPower.32 MR.ReceivedIPower.33 MR.ReceivedIPower.34 MR.ReceivedIPower.35 MR.ReceivedIPower.36 MR.ReceivedIPower.37 MR.ReceivedIPower.38 MR.ReceivedIPower.39 MR.ReceivedIPower.40 MR.ReceivedIPower.41 MR.ReceivedIPower.42 MR.ReceivedIPower.43 MR.ReceivedIPower.44 MR.ReceivedIPower.45 MR.ReceivedIPower.46 MR.ReceivedIPower.47 MR.ReceivedIPower.48 MR.ReceivedIPower.49 MR.ReceivedIPower.50 MR.ReceivedIPower.51 MR.ReceivedIPower.52 '
    ripprb_smr = 'MR.RIPPRB.00 MR.RIPPRB.01 MR.RIPPRB.02 MR.RIPPRB.03 MR.RIPPRB.04 MR.RIPPRB.05 MR.RIPPRB.06 MR.RIPPRB.07 MR.RIPPRB.08 MR.RIPPRB.09 MR.RIPPRB.10 MR.RIPPRB.11 MR.RIPPRB.12 MR.RIPPRB.13 MR.RIPPRB.14 MR.RIPPRB.15 MR.RIPPRB.16 MR.RIPPRB.17 MR.RIPPRB.18 MR.RIPPRB.19 MR.RIPPRB.20 MR.RIPPRB.21 MR.RIPPRB.22 MR.RIPPRB.23 MR.RIPPRB.24 MR.RIPPRB.25 MR.RIPPRB.26 MR.RIPPRB.27 MR.RIPPRB.28 MR.RIPPRB.29 MR.RIPPRB.30 MR.RIPPRB.31 MR.RIPPRB.32 MR.RIPPRB.33 MR.RIPPRB.34 MR.RIPPRB.35 MR.RIPPRB.36 MR.RIPPRB.37 MR.RIPPRB.38 MR.RIPPRB.39 MR.RIPPRB.40 MR.RIPPRB.41 MR.RIPPRB.42 MR.RIPPRB.43 MR.RIPPRB.44 MR.RIPPRB.45 MR.RIPPRB.46 MR.RIPPRB.47 MR.RIPPRB.48 MR.RIPPRB.49 MR.RIPPRB.50 MR.RIPPRB.51 MR.RIPPRB.52 '
    phr_smr = 'MR.PowerHeadRoom.00 MR.PowerHeadRoom.01 MR.PowerHeadRoom.02 MR.PowerHeadRoom.03 MR.PowerHeadRoom.04 MR.PowerHeadRoom.05 MR.PowerHeadRoom.06 MR.PowerHeadRoom.07 MR.PowerHeadRoom.08 MR.PowerHeadRoom.09 MR.PowerHeadRoom.10 MR.PowerHeadRoom.11 MR.PowerHeadRoom.12 MR.PowerHeadRoom.13 MR.PowerHeadRoom.14 MR.PowerHeadRoom.15 MR.PowerHeadRoom.16 MR.PowerHeadRoom.17 MR.PowerHeadRoom.18 MR.PowerHeadRoom.19 MR.PowerHeadRoom.20 MR.PowerHeadRoom.21 MR.PowerHeadRoom.22 MR.PowerHeadRoom.23 MR.PowerHeadRoom.24 MR.PowerHeadRoom.25 MR.PowerHeadRoom.26 MR.PowerHeadRoom.27 MR.PowerHeadRoom.28 MR.PowerHeadRoom.29 MR.PowerHeadRoom.30 MR.PowerHeadRoom.31 MR.PowerHeadRoom.32 MR.PowerHeadRoom.33 MR.PowerHeadRoom.34 MR.PowerHeadRoom.35 MR.PowerHeadRoom.36 MR.PowerHeadRoom.37 MR.PowerHeadRoom.38 MR.PowerHeadRoom.39 MR.PowerHeadRoom.40 MR.PowerHeadRoom.41 MR.PowerHeadRoom.42 MR.PowerHeadRoom.43 MR.PowerHeadRoom.44 MR.PowerHeadRoom.45 MR.PowerHeadRoom.46 MR.PowerHeadRoom.47 MR.PowerHeadRoom.48 MR.PowerHeadRoom.49 MR.PowerHeadRoom.50 MR.PowerHeadRoom.51 MR.PowerHeadRoom.52 MR.PowerHeadRoom.53 MR.PowerHeadRoom.54 MR.PowerHeadRoom.55 MR.PowerHeadRoom.56 MR.PowerHeadRoom.57 MR.PowerHeadRoom.58 MR.PowerHeadRoom.59 MR.PowerHeadRoom.60 MR.PowerHeadRoom.61 MR.PowerHeadRoom.62 MR.PowerHeadRoom.63 '
    sinrul_smr = 'MR.SinrUL.00 MR.SinrUL.01 MR.SinrUL.02 MR.SinrUL.03 MR.SinrUL.04 MR.SinrUL.05 MR.SinrUL.06 MR.SinrUL.07 MR.SinrUL.08 MR.SinrUL.09 MR.SinrUL.10 MR.SinrUL.11 MR.SinrUL.12 MR.SinrUL.13 MR.SinrUL.14 MR.SinrUL.15 MR.SinrUL.16 MR.SinrUL.17 MR.SinrUL.18 MR.SinrUL.19 MR.SinrUL.20 MR.SinrUL.21 MR.SinrUL.22 MR.SinrUL.23 MR.SinrUL.24 MR.SinrUL.25 MR.SinrUL.26 MR.SinrUL.27 MR.SinrUL.28 MR.SinrUL.29 MR.SinrUL.30 MR.SinrUL.31 MR.SinrUL.32 MR.SinrUL.33 MR.SinrUL.34 MR.SinrUL.35 MR.SinrUL.36'

    format_standard_dict = {'MR.RSRP':{'smr':rsrp_smr, 'v':48}, 'MR.RSRQ':{'smr':rsrq_smr, 'v':18}, 'MR.ReceivedIPower':{'smr':rip_smr, 'v':53}, 'MR.RIPPRB':{'smr':ripprb_smr,'v':53}, 'MR.SinrUL':{'smr':sinrul_smr, 'v':37}, 'MR.PowerHeadRoom':{'smr':phr_smr, 'v':64}}
    out_text_list = {}
    try:
        out_mrs_flag_dict = {'MR.RSRP':3, 'MR.RSRQ':3, 'MR.ReceivedIPower':3, 'MR.RIPPRB':3, 'MR.SinrUL':3, 'MR.PowerHeadRoom':3 }
        file_list_mrs = glob.glob(gl.MR_TEST_PATH + '*MRS*.xml')
        for file_name in file_list_mrs:
            out_mrs_flag_dict = {'MR.RSRP':3, 'MR.RSRQ':3, 'MR.ReceivedIPower':3, 'MR.RIPPRB':3, 'MR.SinrUL':3, 'MR.PowerHeadRoom':3 }
            mrs_dom = xml.dom.minidom.parse(file_name)
            mrs_root = mrs_dom.documentElement
            for measurement_entity in mrs_root.getElementsByTagName('measurement'):
                for standard_dict_key in format_standard_dict:
                    if standard_dict_key == measurement_entity.getAttribute('mrName'):
                        #mrName匹配成功
                        out_mrs_flag_dict[standard_dict_key] -= 1
                        #smr数据正确是否
                        smr_list = measurement_entity.getElementsByTagName('smr')
                        if len(smr_list) == 1 and smr_list[0].firstChild.data == format_standard_dict[standard_dict_key]['smr']:
                            out_mrs_flag_dict[standard_dict_key] -= 1
                        #value对应的个数正确是否
                        test_value_num_flag = 0
                        for object_entity in measurement_entity.getElementsByTagName('object'):
                            for value_entity in object_entity.getElementsByTagName('v'):
                                if format_standard_dict[standard_dict_key]['v'] != len(value_entity.firstChild.data.split(' ')) - 1:
                                    test_value_num_flag = 1
                        if test_value_num_flag == 0:
                            out_mrs_flag_dict[standard_dict_key] -= 1
            #判断mrs文件输出是否有问题
            for mrs_dict_key in out_mrs_flag_dict:
                if out_mrs_flag_dict[mrs_dict_key] != 0 and mr_utils.is_mr_item_need_exist(mrs_dict_key) == True:
                    mr_utils.out_text_dict_append_list(out_text_list, file_name, '-[%s not match]' % (mrs_dict_key))
                elif out_mrs_flag_dict[mrs_dict_key] != 3 and mr_utils.is_mr_item_need_exist(mrs_dict_key) == False:
                    mr_utils.out_text_dict_append_list(out_text_list, file_name, '-[%s not match]' % (mrs_dict_key))
            if out_text_list.__contains__(file_name) == True:
                break


        full_name = os.path.join(gl.OUTPUT_PATH, gl.XLS_NAME)
        if int(gl.TEST_CONF['is_57_out_excel']) == 1:
            excel_workbook = openpyxl.load_workbook(full_name)
            work_sheet = excel_workbook.worksheets[0]
            sheet_rows = work_sheet.max_row

            for mrs_dict_key in out_mrs_flag_dict:
                if out_mrs_flag_dict[mrs_dict_key] != 0:
                    for i in range(sheet_row_dict[mrs_dict_key][0], sheet_row_dict[mrs_dict_key][1]+1):
                        work_sheet.cell(i, 20, 'N')
                else:
                    for i in range(sheet_row_dict[mrs_dict_key][0], sheet_row_dict[mrs_dict_key][1]+1):
                        work_sheet.cell(i, 20, 'Y')
            excel_workbook.save(filename=full_name)

        with open(gl.OUT_PATH, 'a') as file_object:
            file_object.write(date_time + " | ")
            file_object.write(gl.TEST_CONF['test_total_time'] + " | " )
            if int(gl.TEST_CONF['is_57_out_excel']) == 1:
                file_object.write('结果已写入->'  )
                file_object.write(gl.XLS_NAME)

            if len(out_text_list) != 0:
                file_object.write('\nerror:\n')
                for file_name in out_text_list:
                    file_object.write('======> %s:'%(file_name))
                    for i in range(len(out_text_list[file_name])):
                        file_object.write('\t%s\n' % (out_text_list[file_name][i]))
            file_object.write('\n')
    except Exception as result:
        raise Exception('<%s>'%(result))

def test58_file_integrity():
    mr_utils.test_out_data_item_header("test_58")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    mro_file_list = glob.glob(gl.MR_TEST_PATH + '*MRO*.xml')

    cellid = int(gl.TEST_CONF['cellid'])
    ideal_sample_num = int(int(gl.MR_CONF['UploadPeriod']) * 60 * 1000 / int(gl.MR_CONF['SamplePeriod']))

    test_cell_id = int(gl.TEST_CONF['cellid']) | int(gl.TEST_CONF['cellid'])  << 8
    out_mro_list = ['MR.LteScRSRP','MR.LteScRSRQ','MR.LteScPHR','MR.LteScSinrUL']
    out_text_list = {}
    xml_mro_item_dict = {"MR.LteScRSRP":2,'MR.LteNcRSRP':3,'MR.LteScRSRQ':4,'MR.LteNcRSRQ':5,'MR.LteScPHR':6, 'MR.LteScRIP':7, 'MR.LteScSinrUL':8 }
    temp_loop_flag = 0
    is_object_empty = 0
    temp_test_flag_dict = {'MR.LteScRSRP':0,'MR.LteScRSRQ':0,'MR.LteScPHR':0,'MR.LteScSinrUL':0,'MR.LteScRIP':0, 'MR.LteNcRSRP':0, 'MR.LteNcRSRQ':0}
    with open(gl.OUT_PATH, 'a') as file_object:
        file_object.write(date_time + " | ")
        file_object.write(gl.TEST_CONF['test_total_time'] + " | " )

    for mro_file in mro_file_list:
        try:
            mro_dom = xml.dom.minidom.parse(mro_file)
            mro_root = mro_dom.documentElement
            mro_item_dict = {}

            for enb_entity in mro_root.getElementsByTagName('eNB'):
                if mr_utils.is_enb_id_exist(enb_entity.getAttribute('id')) == True:
                    for measurement_entity in enb_entity.getElementsByTagName('measurement'):
                        if len(measurement_entity.getElementsByTagName('object')) == 0:
                            is_object_empty = 1
                            if measurement_entity.getElementsByTagName('smr')[0].firstChild.data != 'MR.LteScRIP' and\
                                mr_utils.is_mr_item_need_exist('MR.LteScRSRP') == False and\
                                mr_utils.is_mr_item_need_exist('MR.LteScRSRQ') == False and\
                                mr_utils.is_mr_item_need_exist('MR.LteScPHR') == False and\
                                mr_utils.is_mr_item_need_exist('MR.LteScSinrUL') == False :
                                mr_utils.out_text_dict_append_list(out_text_list, mro_file, 'L3 data None')
                            if measurement_entity.getElementsByTagName('smr')[0].firstChild.data == 'MR.LteScRIP' and\
                                    mr_utils.is_mr_item_need_exist('MR.LteScRIP') == False:
                                mr_utils.out_text_dict_append_list(out_text_list, mro_file, 'L2 data None')
                            break
                        for smr_entity in measurement_entity.getElementsByTagName('smr'):
                            for object_entity in measurement_entity.getElementsByTagName('object'):
                                eci_id = int(object_entity.getAttribute('id').split(':')[0])
                                cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                                object_id = str(cell_id_ret_list[1])
                                object_ue_id = object_id + "|" + object_entity.getAttribute('MmeUeS1apId')
                                if cell_id_ret_list[0] == False:
                                    continue
                                if mro_item_dict.__contains__(object_ue_id) == False  :
                                    mro_item_dict[object_ue_id] = {'MR.LteScRSRP':{'pos':2, 'TimeStamp':0, 'range':[0, 97], 'flag':3, 'num':0, 'item_num':{}}, 'MR.LteScRSRQ':{'pos':3, 'TimeStamp':0, 'range':[0,34],'flag':3, 'num':0, 'item_num':{}},\
                               'MR.LteScPHR':{'pos':4, 'TimeStamp':0, 'range':[0,63],'flag':3, 'num':0, 'item_num':{}}, 'MR.LteScSinrUL':{'pos':5, 'TimeStamp':0, 'range':[0,36],'flag':3, 'num':0, 'item_num':{}}, \
                               'MR.LteScRIP':{'pos':0, 'TimeStamp':0, 'range':[0,511], 'flag':3,'num':0, 'prbnum':{'item_num':{}}}, 'MR.LteNcRSRP':{'pos':8, 'TimeStamp':0, 'range':[0,97],'flag':3, 'num':0, 'item_num':{}},\
                            'MR.LteNcRSRQ':{'pos':9, 'TimeStamp':0, 'range':[0,34], 'flag':3, 'num':0, 'item_num':{} } }
                                mr_utils.get_mr_item_pos(mro_item_dict[object_ue_id], smr_entity.firstChild.data)
                                for mr_name_entity in mro_item_dict[object_ue_id]:
                                    #mr_Name匹配上, flag-1
                                    if re.search(mr_name_entity, smr_entity.firstChild.data) == None:
                                        continue
                                    if mr_name_entity == smr_entity.firstChild.data.split(' ')[mro_item_dict[object_ue_id][mr_name_entity]['pos']]:
                                        mro_item_dict[object_ue_id][mr_name_entity]['flag'] -= 1
                                        if mr_name_entity != 'MR.LteScRIP':
                                            #判断 时间戳是否满足 MR测量的 SamplePeriod
                                            if mro_item_dict[object_ue_id][mr_name_entity]['TimeStamp'] == 0:
                                                mro_item_dict[object_ue_id][mr_name_entity]['TimeStamp'] = mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp'))
                                            else:
                                                time_spec = mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp')) - mro_item_dict[object_ue_id][mr_name_entity]['TimeStamp']
                                                if time_spec != int(gl.MR_CONF['SamplePeriod']):
                                                    temp_test_flag_dict[mr_name_entity] += 1
                                                    mr_utils.out_text_dict_append_list(out_text_list, mro_file, '[{0}]= << TimeStamp duplicate >> TimeStamp:[{1}]\n'.format(mr_name_entity, object_entity.getAttribute('TimeStamp')) \
                                                            if time_spec == 0 else '[{0}]= << TimeStamp gap inaccurate >> TimeStamp:[{1}]\n'.format(mr_name_entity, object_entity.getAttribute('TimeStamp')))
                                                    temp_loop_flag = 1
                                                #当出现时间间隔错误, 继续往下检索
                                                mro_item_dict[object_ue_id][mr_name_entity]['TimeStamp'] = mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp'))

                                        else:

                                            prbnum =  str(mr_utils.is_cell_id_exist(int(object_entity.getAttribute('id').split(':')[0]))[1]) + ':' + object_entity.getAttribute('id').split(':')[2]
                                            if mro_item_dict[object_ue_id][mr_name_entity]['prbnum'].__contains__(prbnum) == False:
                                                mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum] = {'num':0, 'TimeStamp':0}
                                            mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['num'] += 1
                                            #TimeStamp 判断
                                            if mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['TimeStamp'] == 0:
                                                mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['TimeStamp'] = mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp'))
                                            else:
                                                time_spec = mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp')) - mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['TimeStamp']
                                                if time_spec != int(gl.MR_CONF['SamplePeriod']):
                                                    temp_test_flag_dict[mr_name_entity] += 1


                                                    if len(out_text_list[mro_file]) == 0 or re.search(object_entity.getAttribute('TimeStamp'), out_text_list[mro_file][len(out_text_list[mro_file]) - 1]) == None:
                                                        mr_utils.out_text_dict_append_list(out_text_list, mro_file, '[{0}]=<TimeStamp duplicate> TimeStamp:{1}\n'.format(mr_name_entity, object_entity.getAttribute('TimeStamp')) \
                                                          if time_spec == 0 else   '[{0}]=<TimeStamp gap inaccurate> TimeStamp:{1}\n'.format(mr_name_entity, object_entity.getAttribute('TimeStamp')))
                                                    temp_loop_flag = 1
                                            mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['TimeStamp'] = mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp'))

                                        if temp_test_flag_dict[mr_name_entity] == 0:
                                            mro_item_dict[object_ue_id][mr_name_entity]['flag'] -= 1
                                        value_temp_test = 0
                                        for value_entity in object_entity.getElementsByTagName('v'):
                                            value_num = -10
                                            if value_entity.firstChild.data.split(' ')[mro_item_dict[object_ue_id][mr_name_entity]['pos']].isdigit() == True:
                                                value_num = int(value_entity.firstChild.data.split(' ')[mro_item_dict[object_ue_id][mr_name_entity]['pos']])
                                            else:
                                                value_temp_test = 1
                                                continue
                                            if value_num < mro_item_dict[object_ue_id][mr_name_entity]['range'][0] or value_num > mro_item_dict[object_ue_id][mr_name_entity]['range'][1]:
                                                temp_test_flag_dict[mr_name_entity] += 1
                                                mr_utils.out_text_dict_append_list(out_text_list, mro_file, '[' + mr_name_entity + ']={' + 'value Comfusion:' + str(value_num) + '=[' + str(mro_item_dict[object_ue_id][mr_name_entity]['range']) + ']'  + '}\n')
                                                temp_loop_flag = 1
                                                value_temp_test = 1
                                            else:
                                                value_temp_test = 0

                                        mro_item_dict[object_ue_id][mr_name_entity]['num'] += 1

                                        if temp_test_flag_dict[mr_name_entity] == 0:
                                            mro_item_dict[object_ue_id][mr_name_entity]['flag'] -= 1

            with open(gl.OUT_PATH, 'a') as file_object:
                file_object.write('\n [' + mro_file + ']:\n')
                file_object.write('[' + str(cellid) + ']=[')
                file_object.write(str(ideal_sample_num) + 'or' + str(ideal_sample_num+1) + '] |')
                test_integrity_pqsh = 0
                for object_id in mro_item_dict:
                    if re.search(r'NIL', object_id) != None:
                        continue
                    file_object.write( object_id + ":")
                    for mr_name in out_mro_list:
                        file_object.write(str(mro_item_dict[object_id][mr_name]['num']) + ' | ')
                    if mro_item_dict[object_id][mr_name]['num'] != ideal_sample_num and mro_item_dict[object_id][mr_name]['num'] != ideal_sample_num + 1:
                        test_integrity_pqsh = 1
                    if temp_test_flag_dict[mr_name] != 0:
                        test_integrity_pqsh = 1
                file_object.write(str(test_integrity_pqsh == 0) + ' | \n')

                idea_rip_num = ideal_sample_num * len(gl.MR_CONF['SubFrameNum'].split(','))

                file_object.write(date_time + " | ")
                file_object.write(gl.TEST_CONF['test_total_time'] + " | " )
                file_object.write('ideal rip_num=' + str(idea_rip_num) + 'or' + str((ideal_sample_num+1) * len(gl.MR_CONF['SubFrameNum'].split(','))) + ' | ')

                sum_all_prbnum_count = 0
                for object_id in mro_item_dict:
                    if re.search(r'NIL', object_id) == None:
                        continue
                    file_object.write('{')
                    for prbnum_id in mro_item_dict[object_id]['MR.LteScRIP']['prbnum']:
                        if mro_item_dict[object_id]['MR.LteScRIP']['prbnum'][prbnum_id].__contains__('num') == True:
                            file_object.write(' [' + prbnum_id + ']=' + str(mro_item_dict[object_id]['MR.LteScRIP']['prbnum'][prbnum_id]['num'])  )
                            sum_all_prbnum_count += mro_item_dict[object_id]['MR.LteScRIP']['prbnum'][prbnum_id]['num']
                    file_object.write('  [total]=' + str(mro_item_dict[object_id]['MR.LteScRIP']['num']) + ' } | ')

                result_div = sum_all_prbnum_count*1.0 / (idea_rip_num) * 1.0

                rip_accuracy_index = result_div if result_div == 1.0 else sum_all_prbnum_count*1.0 / ((ideal_sample_num + 1) * len(gl.MR_CONF['SubFrameNum'].split(','))) * 1.0
                file_object.write(str( rip_accuracy_index) + ' | \n')
        except  Exception as result:
            raise Exception('[%s]:<%s>'%(mro_file, result))

    with open(gl.OUT_PATH, 'a') as file_object:
        if int(gl.TEST_CONF['is_58_out_excel']) == 1:
            file_object.write('\n结果已写入->'  )
            file_object.write(gl.XLS_NAME)

        if temp_loop_flag == 1:
            file_object.write('\nerror:\n')

            for mr_file in out_text_list:
                file_object.write('=====>' + mr_file + ':\n')
                for i in range(len(out_text_list[mr_file])):
                    file_object.write('\t' + out_text_list[mr_file][i])
        file_object.write('\n')
    if int(gl.TEST_CONF['is_58_out_excel']) == 1:
        full_name = os.path.join(gl.OUTPUT_PATH, gl.XLS_NAME)
        excel_workbook = openpyxl.load_workbook(full_name)
        work_sheet = excel_workbook.worksheets[2]

        for mrs_dict_key in temp_test_flag_dict:
            if temp_test_flag_dict[mrs_dict_key] == 0:
                work_sheet.cell(xml_mro_item_dict[mrs_dict_key], 17, 'Y')
            else:
                work_sheet.cell(xml_mro_item_dict[mrs_dict_key], 17, 'N')
        excel_workbook.save(filename=full_name)



def test59_file_integrity():
    mr_utils.test_out_data_item_header("test_59")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    range_list = [[0,97], [0,97], [0,34], [0,34], [0, 41589], [0, 503], [0, 41589],  [0, 503], [0,1023], [0,63], [0,7], [0,7]]

    out_mre_list = ['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'B1', 'B2']

    mre_smr_head = 'MR.LteScRSRP MR.LteNcRSRP MR.LteScRSRQ MR.LteNcRSRQ MR.LteScEarfcn MR.LteScPci MR.LteNcEarfcn MR.LteNcPci MR.GsmNcellBcch MR.GsmNcellCarrierRSSI MR.GsmNcellNcc MR.GsmNcellBcc'
    mre_smr_list = ['MR.LteScRSRP' ,'MR.LteNcRSRP' ,'MR.LteScRSRQ' ,'MR.LteNcRSRQ' ,'MR.LteScEarfcn' ,'MR.LteScPci' ,'MR.LteNcEarfcn' ,'MR.LteNcPci' ,'MR.GsmNcellBcch' ,'MR.GsmNcellCarrierRSSI' ,'MR.GsmNcellNcc' ,'MR.GsmNcellBcc']

    event_deal_list = []


    xml_out_dict = {'A1':{'item':[2,5], 'event':2}, 'A2':{'item':[6,9], 'event':6}, 'A3':{'item':[10,13], 'event':10}, 'A4':{'item':[14,17], 'event':14}, 'A5':{'item':[18,25], 'event':18}, \
                    'A6':{'item':[26,33], 'event':26}, 'B1':{'item':[34,37], 'event':34}, 'B2':{'item':[38, 45], 'event':38}}

    mre_file_list = glob.glob(gl.MR_TEST_PATH + '*MRE*.xml')

    out_mre_text = {}

    for mre_file_name in mre_file_list:
        try:
            mre_conf_dict = {'A1':{'flag':0, 'pos':[0,2,4,5], 'range':range_list, 'error_pos_list':[]},\
                         'A2':{'flag':0, 'pos':[0,2,4,5], 'range':range_list, 'error_pos_list':[]}, \
                         'A3':{'flag':0, 'pos':[0,1,2,3,4,5,6,7], 'range':range_list, 'error_pos_list':[]},\
                         'A4':{'flag':0, 'pos':[0,1,2,3,4,5,6,7], 'range':range_list, 'error_pos_list':[]},\
                         'A5':{'flag':0, 'pos':[0,1,2,3,4,5,6,7], 'range':range_list, 'error_pos_list':[]},\
                         'A6':{'flag':0, 'pos':[0,1,2,3,4,5,6,7], 'range':range_list, 'error_pos_list':[]},\
                         'B1':{'flag':0, 'pos':[0,2,4,5,8,9,10,11], 'range':range_list, 'error_pos_list':[]},\
                         'B2':{'flag':0, 'pos':[0,2,4,5,8,9,10,11], 'range':range_list, 'error_pos_list':[]}}
            for event_type in mre_conf_dict:
                if re.search(event_type, gl.TEST_CONF['event']) != None:
                    event_deal_list.append(event_type)


            temp_mre_test_flag_dict = {'A1':0, 'A2':0, 'A3':0, 'A4':0, 'A5':0, 'A6':0, 'B1':0, 'B2':0}
            mre_dom = xml.dom.minidom.parse(mre_file_name)
            mre_root = mre_dom.documentElement
            smr_list = mre_root.getElementsByTagName('smr')

            smr_str = smr_list[0].firstChild.data
            if len(smr_list) == 0 :
                temp_flag = 0
                for i in range(len(mre_smr_list)):
                    if mr_utils.is_mr_item_need_exist(mre_smr_list[i]) == True:
                        temp_flag = 1
                        break
                if temp_flag == 1:
                    for event_type in event_deal_list:
                        temp_mre_test_flag_dict[event_type] = 1
                    mr_utils.out_text_dict_append_list(out_mre_text, mre_file_name, 'smr err:' + 'no smr lable' if len(smr_list) == 0 else smr_list[0].firstChild.data)

            if re.search(smr_str, mre_smr_head) == None:
                mr_utils.get_mre_pos_list_by_mapping(mre_conf_dict, smr_str)
                #for event in mre_conf_dict:
                #    print (str(mre_conf_dict[event]['pos']))
            for object_entity in mre_root.getElementsByTagName('object'):
                test_excess_flag = 0
                for event_type in event_deal_list:
                    if event_type == object_entity.getAttribute('EventType'):
                        test_excess_flag = 1
                        mre_conf_dict[event_type]['flag'] = 1
                        for value_entity in object_entity.getElementsByTagName('v'):
                            for pos in range(12):
                                if pos in mre_conf_dict[event_type]['pos']:
                                    value_num = 0
                                    if value_entity.firstChild.data.split(' ')[pos].isdigit() == True:
                                        value_num = int(value_entity.firstChild.data.split(' ')[pos])

                                    if value_num < mre_conf_dict[event_type]['range'][pos][0] or value_num > mre_conf_dict[event_type]['range'][pos][1]:
                                        temp_mre_test_flag_dict[event_type] += 1
                                        mr_utils.out_text_dict_append_list(out_mre_text, mre_file_name, 'value confusion: << {0} = {1} ->({2},{3} )>> event-[{4}] TimeStamp:{5}\n'.\
                                            format(mre_smr_list[pos],str(value_num),str(mre_conf_dict[event_type]['range'][pos][0]), \
                                            str(mre_conf_dict[event_type]['range'][pos][1]), event_type, object_entity.getAttribute('TimeStamp')) )
                                        #print (event_type + '-' + object_entity.getAttribute('TimeStamp'))
                                        if pos not in mre_conf_dict[event_type]['error_pos_list']:
                                            mre_conf_dict[event_type]['error_pos_list'].append(pos)
                                else:
                                    if value_entity.firstChild.data.split(' ')[pos] != 'NIL':
                                        temp_mre_test_flag_dict[event_type] += 1
                                        mr_utils.out_text_dict_append_list(out_mre_text, mre_file_name, 'event format confusion(not NIL):[{0}] event:[{1}}] TimeStamp:[{2}]'.format(mre_smr_list[pos], event_type, object_entity.getAttribute('TimeStamp')))
                                        if pos not in mre_conf_dict[event_type]['error_pos_list']:
                                            mre_conf_dict[event_type]['error_pos_list'].append(pos)
                if test_excess_flag == 0:
                    mr_utils.out_text_dict_append_list(out_mre_text, mre_file_name, 'suplus event:{0} - TimeStamp:{1}\n'.format(object_entity.getAttribute('EventType'), object_entity.getAttribute('TimeStamp')))
            with open(gl.OUT_PATH, 'a') as file_object:
                file_object.write(date_time + " | ")
                file_object.write(gl.TEST_CONF['test_total_time'] + " | " )

                for event_type in out_mre_list:
                    file_object.write('True | ' if event_type in event_deal_list else 'False | ')

                file_object.write(date_time + " | ")
                file_object.write(gl.TEST_CONF['test_total_time'] + " | " )

                for event_type in out_mre_list:
                    if temp_mre_test_flag_dict[event_type] == 0:
                        file_object.write('Y | ' if event_type in event_deal_list else 'N | ' )

                    else:
                        file_object.write('[')
                        for pos in mre_conf_dict[event_type]['error_pos_list']:
                            file_object.write(mre_smr_list[pos] + ' ' )
                        file_object.write('] | ')

                file_object.write('\n\n')
        except Exception as result:
            raise Exception('[%s] <%s>'%(mre_file_name, result))


    if int(gl.TEST_CONF['is_59_out_excel']) == 1:
        full_name = os.path.join(gl.OUTPUT_PATH, gl.XLS_NAME)
        excel_workbook = openpyxl.load_workbook(full_name)
        work_sheet = excel_workbook.worksheets[3]

        for event_type in out_mre_list:
            work_sheet.cell(xml_out_dict[event_type]['event'], 9, 'Y' if event_type in event_deal_list else 'N')
            test_count = 0
            for i in range(xml_out_dict[event_type]['item'][0], xml_out_dict[event_type]['item'][1]+1):
                if mre_conf_dict[event_type]['pos'][test_count] not in mre_conf_dict[event_type]['error_pos_list'] and event_type in event_deal_list :
                    work_sheet.cell(i, 8, 'Y')
                else:
                    work_sheet.cell(i, 8, 'N')
                test_count += 1

        excel_workbook.save(filename=full_name)

    with open(gl.OUT_PATH, 'a') as file_object:
        if int(gl.TEST_CONF['is_59_out_excel']) == 1:
            file_object.write('结果已写入: ' + gl.XLS_NAME)
        for mre_file_name in out_mre_text:
            file_object.write('\n======>' + mre_file_name + ':\n')
            for list_entity in out_mre_text[mre_file_name]:
                file_object.write('\t' + list_entity)




def test61_file_accuracy():
    mr_utils.test_out_data_item_header("test_61")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    try:
        with open(gl.OUT_PATH, 'a') as file_object:
            file_object.write(date_time + " | ")
            file_object.write(gl.TEST_CONF['test_total_time'] + " | " )
    except Exception as result:
        raise Exception('%s'%(result))

def test62_file_accuracy():
    mr_utils.test_out_data_item_header("test_62")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    try:
        with open(gl.OUT_PATH, 'a') as file_object:
            file_object.write(date_time + " | ")
            file_object.write(gl.TEST_CONF['test_total_time'] + " | " )
    except Exception as result:
        raise Exception('%s'%(result))
def test63_file_accuracy():
    mr_utils.test_out_data_item_header("test_63")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)
    try:
        with open(gl.OUT_PATH, 'a') as file_object:
            file_object.write(date_time + " | ")
            file_object.write(gl.TEST_CONF['test_total_time'] + " | " )
    except Exception as result:
        raise Exception('%s'%(result))
def test71_file_accuracy():
    mr_utils.test_out_data_item_header("test_71")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    test_flag = 0
    out_text_list = {}
    try:
        file_list = os.listdir(gl.MR_TEST_PATH)
        re_search_str = gl.TEST_CONF['standard_LTE'] + '_MR'
        for file in file_list:
            full_file = os.path.join(gl.MR_TEST_PATH, file)
            if  os.path.isdir(full_file) == False and re.search(re_search_str, file) != None:
                if mr_utils.MR_xml_file_name_accuracy(file) == False :
                    test_flag |= (0x1 << 1)
                    mr_utils.out_text_dict_append_list(out_text_list, file, 'file_name format err: ->  %s_MR*_%s_%s_%s_YYmmddHHMMSS.xml\n' %( gl.TEST_CONF['standard_LTE'] ,  gl.TEST_CONF['OEM'] , gl.MR_CONF['OmcName'] , gl.TEST_CONF['cellid']))

                if mr_utils.is_mro_correct(full_file)[0] == False:
                    test_flag |= (0x1 << 2)
                    mr_utils.out_text_dict_append_list(out_text_list, file, 'file format  -> %s\n' % (mr_utils.is_mro_correct(full_file)[1]))

                if mr_utils.is_mre_correct(full_file)[0] == False:
                    test_flag |= (0x1 << 2)
                    mr_utils.out_text_dict_append_list(out_text_list, file,'file format  -> %s \n' %(mr_utils.is_mre_correct(full_file)[1] ))
                if mr_utils.is_mrs_correct(full_file)[0] == False:
                    test_flag |= (0x1 << 2)
                    mr_utils.out_text_dict_append_list(out_text_list, file, 'file format  -> %s\n' %(mr_utils.is_mrs_correct(full_file)[1]))

                mr_file_dom = xml.dom.minidom.parse(full_file)
                mr_file_root = mr_file_dom.documentElement
                start_report_Time = mr_utils.get_timestamp_by_str_format(mr_file_root.getElementsByTagName('fileHeader')[0].getAttribute('reportTime'))
                start_report_Time /= 1000
                file_create_time = time.mktime(time.gmtime(os.path.getmtime(full_file)))

                if file_create_time - start_report_Time > int(gl.TEST_CONF['file_delay_time'])*60:
                    test_flag |= (0x1 << 4)
                    mr_utils.out_text_dict_append_list(out_text_list, file, \
                      'file create time:[%s]-[%s] \n' % (str(time.strftime( '%Y-%m-%dT%H:%M:%S',time.localtime(file_create_time))), str(time.strftime( '%Y-%m-%dT%H:%M:%S',time.localtime(start_report_Time)))  ) )

        with open(gl.OUT_PATH, 'a') as file_object:
            file_object.write(date_time + " | ")
            file_object.write(gl.TEST_CONF['test_total_time'] + " | " )
            if test_flag & (0x1 << 1) != 0:
                file_object.write('N | ')
            else:
                file_object.write('Y | ')

            if test_flag & (0x1 << 2) != 0:
                file_object.write('N | ')
            else:
                file_object.write('Y | ')

            if test_flag & (0x1 << 4) != 0:
                file_object.write('[] | N | \n')
            else:
                file_object.write('[] | Y | \n')
            if len(out_text_list) != 0:
                file_object.write('error:\n')
            for file_name in out_text_list:
                file_object.write('======>' + file_name + ' :\n')
                for i in range(len(out_text_list[file_name])):
                    file_object.write('\t' + out_text_list[file_name][i])
    except Exception as result:
        raise Exception('%s'%(result))

def test72_file_accuracy():
    mr_utils.test_out_data_item_header("test_72")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    schema_dict = {"MRO":gl.SOURCE_PATH+"mro_schema.xsd", "MRE":gl.SOURCE_PATH+'mre_schema.xsd', "MRS":gl.SOURCE_PATH+'mrs_schema.xsd'}
    test_flag = 0
    out_text_dict = {}
    file_header_list_name = ['reportTime', 'startTime', 'endTime']

    for time_str in gl.MR_DICT:
        try:
            measureItem_list = {}
            mr_utils.get_measureItem_list(measureItem_list)
            for mr_type in gl.MR_DICT[time_str][0]:
                schema_doc  = etree.parse(schema_dict[mr_type])
                schema_ret = etree.XMLSchema(schema_doc)
                mr_file_full_name = gl.MR_DICT[time_str][0][mr_type]
                if mr_file_full_name == ''  :
                    if mr_utils.is_mr_item_need_exist(mr_type) == True:
                        raise Exception('%s lost %s'%(time_str, mr_type))
                    continue

                data = etree.parse(mr_file_full_name)
                if schema_ret.validate(data) == False:
                    test_flag |= (0x1 << 1)
                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name,schema_ret.error_log )

                if mr_type != 'MRS':
                    mr_doc = xml.dom.minidom.parse(mr_file_full_name)
                    mr_root = mr_doc.documentElement
                    file_header_list = mr_root.getElementsByTagName('fileHeader')
                    enb_id = int(mr_root.getElementsByTagName('eNB')[0].getAttribute('id'))

                    for list_item_name in file_header_list_name:
                        if mr_utils.is_str_format_time(file_header_list[0].getAttribute(list_item_name), gl.TIME_FORMAT) == False:
                            mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, list_item_name + ' format error')
                            test_flag |= (0x1 << 2)
                    if file_header_list[0].getAttribute('period') != '0':
                        mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'upload peroiod error :[0] !=  ' + file_header_list[0].getAttribute('period'))
                        test_flag |= (0x1 << 2)

                    if int(file_header_list[0].getAttribute('jobid')) < 0 or int(file_header_list[0].getAttribute('jobid')) > 4294967295 :
                        mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'jobid :[0-4294967295] !=  ' + file_header_list[0].getAttribute('jobid'))
                        test_flag |= (0x1 << 2)

                    for measurement_entity in mr_root.getElementsByTagName('eNB')[0].getElementsByTagName('measurement'):
                        smr_value = measurement_entity.getElementsByTagName('smr')[0].firstChild.data
                        smr_value_list = smr_value.split(' ')

                        for i in range(len(smr_value_list)):
                            temp_flag = 0
                            for mr_item in measureItem_list:
                                if smr_value_list[i] == mr_item:
                                    temp_flag = 1
                                    measureItem_list[mr_item] += 1
                                    break
                            if temp_flag == 0:
                                mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'surplus the mr_item : %s'%(smr_value_list[i]))
                        for object_entity in measurement_entity.getElementsByTagName('object'):
                            eci_id = int(object_entity.getAttribute('id').split(':')[0])
                            enb_id = eci_id >> 8 & 0xff
                            cell_id_ret_list = mr_utils.is_cell_id_exist(eci_id)
                            if cell_id_ret_list[0] == False or mr_utils.is_enb_id_exist(enb_id) == False:
                                mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '[%s] <cell_id(%d) or enb_id(%d) not exist>'%(\
                                    smr_value, cell_id_ret_list[1], enb_id))
                                test_flag |= (0x1 << 3)
                            if mr_utils.is_str_format_time(object_entity.getAttribute('TimeStamp'), gl.TIME_FORMAT) == False:
                                mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '[%s] <timestamp format error> %s'%(\
                                   smr_value, object_entity.getAttribute('TimeStamp')))
                                test_flag |= (0x1 << 3)
                            if smr_value_list[0] == 'MR.LteScRIP':
                                if  mr_utils.is_eci_correct(int(object_entity.getAttribute('id').split(':')[0])) == False or \
                                    re.search(object_entity.getAttribute('id').split(':')[2], gl.MR_CONF['SubFrameNum']) == None or\
                                    object_entity.getAttribute('MmeUeS1apId') != 'NIL' or \
                                    object_entity.getAttribute('MmeGroupId') != 'NIL' or \
                                    object_entity.getAttribute('MmeCode') != 'NIL' :
                                        test_flag |= (0x1 << 4)
                                        mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'MR.LteScRIP:<rip object_format error> %s ' % (object_entity.getAttribute('TimeStamp')))
                            else:

                                if mr_utils.is_eci_correct(int(object_entity.getAttribute('id').split(':')[0])) == False :
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '[L3] <eci_id not match> %s'% (object_entity.getAttribute('TimeStamp')))
                                    test_flag |= (0x1 << 3)
                            temp_value_flag = 0

                            for value_entity in object_entity.getElementsByTagName('v'):
                                value_str = value_entity.firstChild.data
                                if mr_utils.is_mr_value_correct(smr_value, value_str) == False:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'value not match')
                                    test_flag |= (0x1 << 4)
                                    temp_value_flag = 1
                            if temp_value_flag == 0:
                                for mr_entity in measureItem_list:
                                    if mr_entity != 'MR.RSRP' or mr_entity != 'MR.RSRQ' or mr_entity != 'MR.ReceivedIPower' or mr_entity != 'MR.RIPPRB' or mr_entity != 'MR.PowerHeadRoom':
                                        measureItem_list[mr_entity] += 1

                else:
                    mr_doc = xml.dom.minidom.parse(mr_file_full_name)
                    mr_root = mr_doc.documentElement
                    file_header_list = mr_root.getElementsByTagName('fileHeader')
                    #得到MRO的earfcn，判断MRS文件中的earfcn是否一致
                    mro_doc = xml.dom.minidom.parse(gl.MR_DICT[time_str][0]['MRO'])
                    mro_root = mro_doc.documentElement
                    earfcn_value = mro_root.getElementsByTagName('measurement')[0].getElementsByTagName('object')[0].getElementsByTagName('v')[0].firstChild.data.split(' ')[0] if \
                        len(mro_root.getElementsByTagName('measurement')[0].getElementsByTagName('object')) != 0 else '-1'

                    #获得子帧分帧
                    prbnum_list = []
                    if re.search(r',', gl.MR_CONF['PrbNum']) == None:
                        for i in range(int(gl.MR_CONF['PrbNum'].split('....')[0]), 1 + int(gl.MR_CONF['PrbNum'].split('....')[1])):
                            prbnum_list.append(str(i))
                    else:
                        prbnum_list = gl.MR_CONF['PrbNum'].split(',')
                    subfram_list = gl.MR_CONF['SubFrameNum'].split(',')

                    for list_item_name in file_header_list_name:
                        if mr_utils.is_str_format_time(file_header_list[0].getAttribute(list_item_name), gl.TIME_FORMAT) == False:
                            mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '%s format error'%(list_item_name))
                            test_flag |= (0x1 << 2)

                    if file_header_list[0].getAttribute('period') != gl.MR_CONF['UploadPeriod']:
                        mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'upload peroiod error :[0] !=  ' + file_header_list[0].getAttribute('period'))
                        test_flag |= (0x1 << 2)

                    if int(file_header_list[0].getAttribute('jobid')) < 0 or int(file_header_list[0].getAttribute('jobid')) > 4294967295 :
                        mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, 'jobid :[0-4294967295] !=  ' + file_header_list[0].getAttribute('jobid'))
                        test_flag |= (0x1 << 2)

                    for measurement_entity in mr_root.getElementsByTagName('eNB')[0].getElementsByTagName('measurement'):
                        mr_name = measurement_entity.getAttribute('mrName')
                        if measureItem_list.__contains__(mr_name) == True:
                            measureItem_list[mr_name] += 1
                        else:
                            mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '<surplus mrItem> %s'%(mr_name))
                        smr_value = measurement_entity.getElementsByTagName('smr')[0].firstChild.data
                        object_list = measurement_entity.getElementsByTagName('object')

                        if mr_name == "MR.ReceivedIPower":
                            if len(subfram_list)*len(gl.TEST_CONF['cellid'].split(',')) != len(object_list):
                                mr_utils.string_to_list(out_text_dict, mr_file_full_name, "MR.ReceivedIPower:object format error:[%s]" % (str(subfram_list[0:len(subfram_list)])))
                                test_flag |= (0x1 << 2)
                            for object_entity in object_list:
                                id_str = object_entity.getAttribute('id')
                                id_list = id_str.split(':')
                                if mr_utils.is_eci_correct(int(id_list[0])) == False or (id_list[1] != earfcn_value and earfcn_value != '-1') or id_list[2] not in subfram_list:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name,"MR.ReceivedIPower:object format error:[%s]"%(id_str))
                                    test_flag |= (0x1 << 4)

                        elif mr_name == 'MR.RIPPRB':
                            if len(subfram_list)*len(prbnum_list)*len(gl.TEST_CONF['cellid'].split(',')) != len(object_list):
                                mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, "MR.RIPPRB:object format error:[%s * %s]" % (str(subfram_list[0:len(subfram_list)]) , str(prbnum_list[0:len(prbnum_list)])))
                                test_flag |= (0x1 << 4)
                            for object_entity in object_list:
                                id_str = object_entity.getAttribute('id')
                                id_list = id_str.split(':')
                                if mr_utils.is_eci_correct(int(id_list[0])) == False or (id_list[1] != earfcn_value and earfcn_value != '-1') or id_list[2] not in subfram_list or id_list[2] not in prbnum_list:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, "MR.RIPPRB:object format error:[%s]"%(id_str))
                                    test_flag |= (0x1 << 4)
                        else:
                            if len(object_list) != len(gl.TEST_CONF['cellid'].split(',')):
                                mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '%s: object format error [num not match] - [%d]'%(mr_name,len(object_list)))
                                test_flag |= (0x1 << 4)
                            for object_entity in object_list:
                                id_str = object_entity.getAttribute('id')
                                id_list = id_str.split(':')
                                if mr_utils.is_eci_correct(int(id_list[0])) == False:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '%s <object id not match > %s'%(mr_name, id_str))
                                    test_flag |= (0x1 << 4)
                        temp_value_flag = 0

                        for object_entity in object_list:
                            for value_entity in object_entity.getElementsByTagName('v'):
                                value_str = value_entity.firstChild.data
                                if mr_utils.is_mrs_measurement_smr_value_correct(mr_name, smr_value, value_str) == False:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mr_file_full_name, '<smr or value not match> \n--->smr:%s \n---> value:%s'%(smr_value, value_str))
                                    test_flag |= (0x1 << 4)
                                    temp_value_flag = 1
                        if temp_value_flag == 0:
                            measureItem_list[mr_name] += 1

        except Exception as result:
            raise Exception('<%s> MRO:%s MRE:%s MRS:%s'%(result, gl.MR_DICT[time_str][0]['MRO'], gl.MR_DICT[time_str][0]['MRE'], gl.MR_DICT[time_str][0]['MRS']))

    with open(gl.OUT_PATH, 'a') as file_object:
        file_object.write(date_time + " | ")
        file_object.write(gl.TEST_CONF['test_total_time'] + " | " )

        file_object.write('N | N | ' if test_flag & (0x1 << 1) != 0 else 'Y | Y | ')
        file_object.write('N | N | ' if test_flag & (0x1 << 2) != 0 else 'Y | Y | ' )
        file_object.write('N | ' if test_flag & (0x1 << 3) != 0 else 'Y | ')
        file_object.write('N | \n' if test_flag & (0x1 << 4) != 0 else 'Y | \n')
        if test_flag != 0:
            for file_name in out_text_dict:
                file_object.write('=====>' + file_name + ' :\n')
                for i in range(len(out_text_dict[file_name])):
                    file_object.write('\t' + str(out_text_dict[file_name][i]) + '\n')

def test73_file_accuracy():
    mr_utils.test_out_data_item_header("test_73")
    date_time = mr_utils.get_time_format(gl.TIME_OUTPUT_FORMAT)

    mro_file_list = glob.glob(gl.MR_TEST_PATH + '*MRO*.xml')
    max_mro_object_num = math.ceil( (int(gl.MR_CONF['UploadPeriod']) * 60 * 1000) / float(int(gl.MR_CONF['SamplePeriod'])))
    mro_file_flag_dict = {}
    out_text_dict = {}
    for mro_file in mro_file_list:
        try:
            mro_dom = xml.dom.minidom.parse(mro_file)
            mro_root = mro_dom.documentElement
            time_stamp_tmep = 0
            if mro_file_flag_dict.__contains__(mro_file) == False:
                mro_file_flag_dict[mro_file] = {}
            for measurement_entity in mro_root.getElementsByTagName('measurement'):
                if measurement_entity.getElementsByTagName('smr')[0].firstChild.data.split(' ')[0] == 'MR.LteScRIP' :
                    continue
                for object_entity in measurement_entity.getElementsByTagName('object'):
                    ue_mme_name = object_entity.getAttribute('MmeUeS1apId') + '|' + object_entity.getAttribute('MmeGroupId') + '|' + object_entity.getAttribute('MmeCode')
                    if mro_file_flag_dict[mro_file].__contains__(ue_mme_name) == False:
                        mro_file_flag_dict[mro_file][ue_mme_name] = {'mro_object_num':0, 'is_ascend':0}
                    if time_stamp_tmep != 0:
                        time_spec = mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp')) - time_stamp_tmep
                        if time_spec != int(gl.MR_CONF['SamplePeriod']):
                            mro_file_flag_dict[mro_file][ue_mme_name]['is_ascend'] = 1
                            mr_utils.out_text_dict_append_list(out_text_dict, mro_file, "time spec: [%d]-[%d]=[%d] - [%d] "%(mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp')), time_stamp_tmep, time_spec, int(gl.MR_CONF['SamplePeriod'])))
                    time_stamp_tmep = mr_utils.get_timestamp_by_str_format(object_entity.getAttribute('TimeStamp'))
                    mro_file_flag_dict[mro_file][ue_mme_name]['mro_object_num'] += 1
        except Exception as result:
            raise Exception('%s <%s>'%(mro_file, result))

    with open(gl.OUT_PATH, 'a') as file_object:
        file_object.write(date_time + " | ")
        file_object.write(gl.TEST_CONF['test_total_time'] + " | " )
        file_object.write(str(max_mro_object_num) + ' | \n')
        for mro_file in mro_file_flag_dict:
            file_object.write('======>' + mro_file + ": \n\t")
            if len(mro_file_flag_dict[mro_file]) == 0:
                file_object.write("no L3 information\n")
                continue
            for ue_mme_name in mro_file_flag_dict[mro_file]:
                file_object.write( "[" + ue_mme_name + "]: [" + str(mro_file_flag_dict[mro_file][ue_mme_name]['mro_object_num']) + "]-["+ str(mro_file_flag_dict[mro_file][ue_mme_name]['is_ascend'] == 0) + "]  ")
            if mro_file_flag_dict[mro_file][ue_mme_name]['mro_object_num'] > max_mro_object_num or mro_file_flag_dict[mro_file][ue_mme_name]['is_ascend'] != 0:
                file_object.write(' | N |\n')
            else:
                file_object.write(' | Y |\n')

def test_add_timestamp_number():
    with open(gl.OUT_PATH, 'a') as file_object:
        file_object.write('\n<-------------------------->\n')
        file_object.write('test the timestamp [l2-l3]\nresult:\n')

    mro_file_list = glob.glob(gl.MR_TEST_PATH + '*MRO*.xml')
    out_text_dict = {}
    out_list = ['l3','l2']
    smr_l3_str = 'MR.LteScEarfcn MR.LteScPci MR.LteScRSRP MR.LteScRSRQ MR.LteScPHR MR.LteScSinrUL MR.LteNcEarfcn MR.LteNcPci MR.LteNcRSRP MR.LteNcRSRQ'
    smr_rip_str = 'MR.LteScRIP'
    for mro_file in mro_file_list:
        try:
            mro_doc = xml.dom.minidom.parse(mro_file)
            mro_root = mro_doc.documentElement
            timestamp_dict = {'l3': {}, 'l2':{}}
            l3_timestamp_temp = 0
            l2_timestamp_temp = 0

            for enb_entity in mro_root.getElementsByTagName('eNB'):
                start_time_stamp = mr_utils.get_timestamp_by_str_format(mro_root.getElementsByTagName('fileHeader')[0].getAttribute('startTime')) + int(gl.MR_CONF['SamplePeriod'])

                end_time_stamp = mr_utils.get_timestamp_by_str_format(mro_root.getElementsByTagName('fileHeader')[0].getAttribute('endTime'))
                measurement_list = enb_entity.getElementsByTagName('measurement')
                if len(measurement_list) != 2 :
                    continue
                l3_smr_list = measurement_list[0].getElementsByTagName('smr')
                l3_object_list = measurement_list[0].getElementsByTagName('object')
                l2_smr_list = measurement_list[1].getElementsByTagName('smr')
                l2_object_list = measurement_list[1].getElementsByTagName('object')
                if len(l3_smr_list) != 1 or len(l3_object_list) == 0 or len(l2_smr_list) != 1 or len(l2_object_list) == 0:
                    continue
                if re.search(smr_l3_str, l3_smr_list[0].firstChild.data) == None:
                    continue
                if re.search(smr_rip_str, l2_smr_list[0].firstChild.data) == None:
                    continue
                for l3_object_entity in l3_object_list:
                    cell_id_ret_list = mr_utils.is_cell_id_exist(int(l3_object_entity.getAttribute('id').split(':')[0]))
                    cell_id = str(cell_id_ret_list[1])
                    if timestamp_dict['l3'].__contains__(cell_id) == False:
                        timestamp_dict['l3'][cell_id] = {'num':0, 'time_str':[], 'time_stamp':[]}

                    time_str = l3_object_entity.getAttribute('TimeStamp')
                    time_stamp = mr_utils.get_timestamp_by_str_format(time_str)

                    if l3_timestamp_temp != time_stamp:
                        timestamp_dict['l3'][cell_id]['num'] += 1
                        l3_timestamp_temp = time_stamp
                        timestamp_dict['l3'][cell_id]['time_str'].append(time_str)
                        timestamp_dict['l3'][cell_id]['time_stamp'].append(time_stamp)

                for l2_object_entity in l2_object_list:
                    cell_id_ret_list = mr_utils.is_cell_id_exist(int(l2_object_entity.getAttribute('id').split(':')[0]))
                    cell_id = str(cell_id_ret_list[1])
                    if timestamp_dict['l2'].__contains__(cell_id) == False:
                        timestamp_dict['l2'][cell_id] = {'num':0, 'time_str':[], 'time_stamp':[]}
                    time_str = l2_object_entity.getAttribute('TimeStamp')
                    time_stamp = mr_utils.get_timestamp_by_str_format(time_str)

                    if l2_timestamp_temp != time_stamp:
                        timestamp_dict['l2'][cell_id]['num'] += 1
                        l2_timestamp_temp = time_stamp
                        timestamp_dict['l2'][cell_id]['time_str'].append(time_str)
                        timestamp_dict['l2'][cell_id]['time_stamp'].append(time_stamp)
                for cell_id in timestamp_dict['l2'] if len(timestamp_dict['l2']) >= len(timestamp_dict['l3']) else timestamp_dict['l3']:
                    if timestamp_dict['l2'].__contains__(cell_id) == True and timestamp_dict['l3'].__contains__(cell_id) == True:
                        if timestamp_dict['l2'][cell_id]['num'] != timestamp_dict['l3'][cell_id]['num']:
                            mr_utils.out_text_dict_append_list(out_text_dict, mro_file, '[cellid:%s] object_num : [l3]-%d != [l2]-%d \n'%(cell_id, timestamp_dict['l3'][cell_id]['num'], timestamp_dict['l2'][cell_id]['num']))

                for l_entity in out_list:
                    for cell_id in timestamp_dict[l_entity]:
                        if timestamp_dict[l_entity][cell_id]['time_stamp'][0] != start_time_stamp:
                            mr_utils.out_text_dict_append_list(out_text_dict, mro_file, '[%s] <%s startTimeStamp not match>: %s(start) != %s(first) \n'%(cell_id,l_entity,\
                                mr_utils.get_time_format_by_timestamp(start_time_stamp),\
                                timestamp_dict[l_entity]['time_str'][0]))
                        if timestamp_dict[l_entity][cell_id]['time_stamp'][len(timestamp_dict[l_entity][cell_id]['time_stamp']) - 1] != end_time_stamp:
                            mr_utils.out_text_dict_append_list(out_text_dict, mro_file, '[%s] <%s endTimeStamp not match>: %s(end) != %s(last) \n'%(cell_id,l_entity,\
                                mro_root.getElementsByTagName('fileHeader')[0].getAttribute('endTime'),\
                                timestamp_dict[l_entity]['time_str'][len(timestamp_dict[l_entity]['time_stamp']) - 1]))
                for cell_id in timestamp_dict['l2'] if len(timestamp_dict['l2']) >= len(timestamp_dict['l3']) else timestamp_dict['l3']:
                    if timestamp_dict['l2'].__contains__(cell_id) == True and timestamp_dict['l3'].__contains__(cell_id) == True:
                        l3_index = 0
                        l2_index = 0
                        suplus_temp = 0
                        for i in range(0, timestamp_dict['l3'][cell_id]['num'] if timestamp_dict['l3'][cell_id]['num'] >= timestamp_dict['l2'][cell_id]['num'] else timestamp_dict['l2'][cell_id]['num'] ):
                            if l2_index > timestamp_dict['l2'][cell_id]['num']-1:
                                l2_index = timestamp_dict['l2'][cell_id]['num']-1
                                suplus_temp += 1
                            if l3_index > timestamp_dict['l3'][cell_id]['num']-1:
                                l3_index = timestamp_dict['l3'][cell_id]['num']-1
                                suplus_temp += 1
                            if suplus_temp > 3:
                                if l3_index == timestamp_dict['l3'][cell_id]['num']-1 and l2_index != timestamp_dict['l2'][cell_id]['num']-1:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mro_file, '[{4}]<l3 lost part of data>: num:[l3={0}]--[l2={1}] TimeStamp:[l3={2}--l2={3}]'.format(timestamp_dict['l3'][cell_id]['num']-1,\
                                                                                timestamp_dict['l2'][cell_id]['num']-1, timestamp_dict['l3'][cell_id]['time_str'][l3_index], timestamp_dict['l2'][cell_id]['time_str'][len(timestamp_dict['l2'][cell_id]['time_str']) - 1], cell_id))
                                elif l3_index != timestamp_dict['l3'][cell_id]['num']-1 and l2_index == timestamp_dict['l2'][cell_id]['num']-1:
                                    mr_utils.out_text_dict_append_list(out_text_dict, mro_file,'[{4}]<l2 lost part of data: num>:[l3={0}]--[l2={1}] TimeStamp:[l3={2}--l2={3}]'.format(timestamp_dict['l3'][cell_id]['num']-1,\
                                                                                timestamp_dict['l2'][cell_id]['num']-1, timestamp_dict['l3'][cell_id]['time_str'][len(timestamp_dict['l3'][cell_id]['time_str']) - 1], timestamp_dict['l2'][cell_id]['time_str'][l2_index], cell_id))
                                break

                            if timestamp_dict['l3'][cell_id]['time_stamp'][l3_index] == timestamp_dict['l2'][cell_id]['time_stamp'][l2_index]:
                                l2_index += 1
                                l3_index += 1
                            elif timestamp_dict['l3'][cell_id]['time_stamp'][l3_index] < timestamp_dict['l2'][cell_id]['time_stamp'][l2_index]:
                                mr_utils.out_text_dict_append_list(out_text_dict, mro_file, '[%s] <timestamp not match>: [l3]-[%d]-%s < [l2]-[%d]-%s \n'%(cell_id,l3_index,timestamp_dict['l3']['time_str'][l3_index], l3_index,timestamp_dict['l2']['time_str'][l2_index] ))
                                l3_index += 1
                            elif timestamp_dict['l3'][cell_id]['time_stamp'][l3_index] > timestamp_dict['l2'][cell_id]['time_stamp'][l2_index]:
                                mr_utils.out_text_dict_append_list(out_text_dict, mro_file, '[%s] <timestamp not match>: [l3]-[%d]-%s > [l2]-[%d]-%s \n'%(cell_id,l3_index,timestamp_dict['l3']['time_str'][l3_index], l3_index,timestamp_dict['l2']['time_str'][l2_index] ))
                                l2_index += 1
        except Exception as result:
            raise Exception('%s <%s>'%(mro_file, result))

    with open(gl.OUT_PATH, 'a') as file_object:
        file_object.write('test ok\n')
        if len(out_text_dict) != 0:
            file_object.write('error:\n')
        for mro_file_name in out_text_dict:
            file_object.write('=====>  ' + mro_file_name + ':\n')
            for i in range(0, len(out_text_dict[mro_file_name])):
                file_object.write('\t' + out_text_dict[mro_file_name][i] )

def test_add_mro_s_mapping():

    #范围
    rsrp_range_list = []
    rsrq_range_list = []
    phr_range_list = []
    sinrul_range_list = []
    rip_range_list = []

    rsrp_range_list.append(20)
    for i in range(25, 60):
        rsrp_range_list.append(i)
    for i in range(60, 81, 2):
        rsrp_range_list.append(i)
    rsrp_range_list.append(97)
    for i in range(0, 35, 2):
        rsrq_range_list.append(i)
    for i in range(64):
        phr_range_list.append(i)
    for i in range(0, 37):
        sinrul_range_list.append(i)

    for i in range(0, 521, 10):
        rip_range_list.append(i)

    out_mro_list = ['MR.LteScRSRP','MR.LteScRSRQ','MR.LteScPHR','MR.LteScSinrUL']
    out_text_list = {}
    for time_entity in gl.MR_DICT:


        cellid_item_dict = {'MR.LteScRSRP':{'pos':2, 'TimeStamp':0, 'range':[0, 97], 'flag':3, 'num':0}, 'MR.LteScRSRQ':{'pos':3, 'TimeStamp':0, 'range':[0,34],'flag':3, 'num':0},\
                               'MR.LteScPHR':{'pos':4, 'TimeStamp':0, 'range':[0,63],'flag':3, 'num':0}, 'MR.LteScSinrUL':{'pos':5, 'TimeStamp':0, 'range':[0,36],'flag':3, 'num':0}, \
                               'MR.LteScRIP':{'pos':0, 'TimeStamp':0, 'range':[0,511], 'flag':3,'num':0, 'prbnum':{}}, 'MR.LteNcRSRP':{'pos':8, 'TimeStamp':0, 'range':[0,97],'flag':3, 'num':0},\
                            'MR.LteNcRSRQ':{'pos':9, 'TimeStamp':0, 'range':[0,34], 'flag':3, 'num':0 } }
        item_range_list = {'MR.LteScRSRP':rsrp_range_list,
                           'MR.LteScRSRQ':rsrq_range_list,
                           'MR.LteScPHR':phr_range_list,
                           'MR.LteScSinrUL':sinrul_range_list,
                           'MR.LteScRIP':rip_range_list}
        try:

            temp_loop_flag = 0
            is_object_empty = 0
            mro_item_dict = {}
            temp_test_flag_dict = {'MR.LteScRSRP':0,'MR.LteScRSRQ':0,'MR.LteScPHR':0,'MR.LteScSinrUL':0,'MR.LteScRIP':0, 'MR.LteNcRSRP':0, 'MR.LteNcRSRQ':0}
            mro_measurement_list = []
            mrs_measurement_list = []
            if re.search(r'MRO', gl.MR_CONF['MeasureType']) != None:
                mro_file_name = gl.MR_DICT[time_entity][0]['MRO']
                mro_dom = xml.dom.minidom.parse(mro_file_name)
                mro_root = mro_dom.documentElement
                mro_measurement_list = mro_root.getElementsByTagName('measurement')
            if re.search(r'MRS', gl.MR_CONF['MeasureType']) != None:
                mrs_file_name = gl.MR_DICT[time_entity][0]['MRS']
                mrs_dom = xml.dom.minidom.parse(mrs_file_name)
                mrs_root = mrs_dom.documentElement
                mrs_measurement_list = mrs_root.getElementsByTagName('measurement')

            for measurement_entity in mro_measurement_list:
                if len(measurement_entity.getElementsByTagName('object')) == 0:
                    is_object_empty = 1
                    continue
                for smr_entity in measurement_entity.getElementsByTagName('smr'):
                    for object_entity in measurement_entity.getElementsByTagName('object'):
                        object_ue_id = object_entity.getAttribute('id').split(':')[0]
                        if mro_item_dict.__contains__(object_ue_id) == False :
                            mro_item_dict[object_ue_id] = {'MR.LteScRSRP':{'pos':2, 'mrs_item':'MR.RSRP', 'range':[0, 97], 'flag':3, 'num':0, 'item_num':{}}, 'MR.LteScRSRQ':{'pos':3, 'mrs_item':'MR.RSRQ', 'range':[0,34],'flag':3, 'num':0, 'item_num':{}},\
                               'MR.LteScPHR':{'pos':4, 'mrs_item':'MR.PowerHeadRoom', 'range':[0,63],'flag':3, 'num':0, 'item_num':{}}, 'MR.LteScSinrUL':{'pos':5, 'mrs_item':'MR.SinrUL', 'range':[0,36],'flag':3, 'num':0, 'item_num':{}}, \
                               'MR.LteScRIP':{'pos':0, 'mrs_item':'MR.ReceivedIPower', 'range':[0,511], 'flag':3,'num':0, 'prbnum':{'item_num':{}}}, 'MR.LteNcRSRP':{'pos':8, 'mrs_item':0, 'range':[0,97],'flag':3, 'num':0, 'item_num':{}},\
                            'MR.LteNcRSRQ':{'pos':9, 'mrs_item':0, 'range':[0,34], 'flag':3, 'num':0, 'item_num':{} } }
                            mr_utils.get_mro_pos_list_by_mapping(mro_item_dict[object_ue_id], smr_entity.firstChild.data)
                        for mr_name_entity in mro_item_dict[object_ue_id]:
                            #mr_Name匹配上, flag-1
                            if re.search(mr_name_entity, smr_entity.firstChild.data) == None or mro_item_dict[object_ue_id][mr_name_entity]['pos'] == -1:
                                continue
                            if mr_name_entity == smr_entity.firstChild.data.split(' ')[mro_item_dict[object_ue_id][mr_name_entity]['pos']]:
                                if mr_name_entity != 'MR.LteScRIP':
                                    #TODO: 得到value, 然后对应的加上
                                    if len(object_entity.getElementsByTagName('v')) != 0:
                                        value_str = object_entity.getElementsByTagName('v')[0].firstChild.data.split(' ')[mro_item_dict[object_ue_id][mr_name_entity]['pos']]
                                        if value_str.isdigit() == True:
                                            value = int(value_str)
                                            for idx in range(len(item_range_list[mr_name_entity])):
                                                if mro_item_dict[object_ue_id][mr_name_entity].__contains__('item_num') == False:
                                                    mro_item_dict[object_ue_id][mr_name_entity]['item_num'] = {}
                                                if mro_item_dict[object_ue_id][mr_name_entity]['item_num'].__contains__(idx) == False:
                                                    mro_item_dict[object_ue_id][mr_name_entity]['item_num'][idx] = 0
                                                if value <= item_range_list[mr_name_entity][idx]:
                                                    mro_item_dict[object_ue_id][mr_name_entity]['item_num'][idx] += 1
                                                    break
                                else:
                                    prbnum = object_entity.getAttribute('id').split(':')[0] + ':' + object_entity.getAttribute('id').split(':')[2]
                                    if mro_item_dict[object_ue_id][mr_name_entity]['prbnum'].__contains__(prbnum) == False:
                                        mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum] = {'num':0, 'TimeStamp':0, 'item_num':{}}
                                    #TODO:获得value, 加上
                                    if len(object_entity.getElementsByTagName('v')) != 0:
                                        value_str = object_entity.getElementsByTagName('v')[0].firstChild.data.split(' ')[mro_item_dict[object_ue_id][mr_name_entity]['pos']]
                                        if value_str.isdigit() == True:
                                            value = int(value_str)
                                            for idx in range(len(item_range_list[mr_name_entity])):
                                                if mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum].__contains__('item_num') == False:
                                                    mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['item_num'] = {}
                                                if mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['item_num'].__contains__(idx) == False:
                                                    mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['item_num'][idx] = 0
                                                if value <= item_range_list[mr_name_entity][idx]:
                                                    mro_item_dict[object_ue_id][mr_name_entity]['prbnum'][prbnum]['item_num'][idx] += 1
                                                    break

            for measurement_entity in mrs_measurement_list:
                for object_ue_id in mro_item_dict:
                    for mr_name in mro_item_dict[object_ue_id]:
                        if mro_item_dict[object_ue_id][mr_name]['mrs_item'] == measurement_entity.getAttribute('mrName'):
                            if mr_name != 'MR.LteScRIP':
                                for object_entity in measurement_entity.getElementsByTagName('object'):
                                    if object_entity.getAttribute('id') == object_ue_id :
                                        value_entity_list = object_entity.getElementsByTagName('v')[0].firstChild.data.strip().split(' ')
                                        for i in range(len(item_range_list[mr_name])):
                                            if (int(value_entity_list[i]) != 0 and mro_item_dict[object_ue_id][mr_name]['item_num'].__contains__(i) == False) or \
                                                    (mro_item_dict[object_ue_id][mr_name]['item_num'].__contains__(i) == True and \
                                                     int(value_entity_list[i]) != mro_item_dict[object_ue_id][mr_name]['item_num'][i]) :
                                                mr_utils.out_text_dict_append_list(out_text_list, mrs_file_name, '[%s]-[%d]=[%d(mro) != %d(mrs)]'%(mr_name, i,
                                                    mro_item_dict[object_ue_id][mr_name]['item_num'][i] if mro_item_dict[object_ue_id][mr_name]['item_num'] else 0,\
                                                    int(value_entity_list[i])))
                            else:
                                for object_entity in measurement_entity.getElementsByTagName('object'):
                                    prbnum = object_entity.getAttribute('id').split(':')[0] + ':' + object_entity.getAttribute('id').split(':')[2]
                                    if object_entity.getAttribute('id').split(':')[0] == object_ue_id :
                                        for prb_entity in mro_item_dict[object_ue_id][mr_name]['prbnum']:
                                            if prb_entity == prbnum:
                                                value_entity_list = object_entity.getElementsByTagName('v')[0].firstChild.data.split(' ')
                                                for i in range(len(item_range_list[mr_name])):
                                                    if (int(value_entity_list[i]) != 0 and mro_item_dict[object_ue_id][mr_name]['prbnum'][prb_entity]['item_num'].__contains__(i) == False) or \
                                                        (mro_item_dict[object_ue_id][mr_name]['prbnum'][prb_entity]['item_num'].__contains__(i) == True and int(value_entity_list[i]) != mro_item_dict[object_ue_id][mr_name]['prbnum'][prb_entity]['item_num'][i]):
                                                        mr_utils.out_text_dict_append_list(out_text_list, mrs_file_name, '[%s]-[prb:%s id:%d]=[%d(mro) != %d(mrs)]'%(mr_name, prb_entity, i,
                                                            mro_item_dict[object_ue_id][mr_name]['prbnum'][prb_entity]['item_num'][i] if mro_item_dict[object_ue_id][mr_name]['prbnum'][prb_entity]['item_num'].__contains__(i) == True else 0, int(value_entity_list[i])))
        except Exception as result:
            raise Exception('<%s> MRO:%s MRS:%s'%(result, gl.MR_DICT[time_entity][0]['MRO'], gl.MR_DICT[time_entity][0]['MRS']))
    with open(gl.OUT_PATH, 'a') as file_object:
        file_object.write('\n [' + ' add mro-mrs num mapping ' + ']:\n')
        temp_flag = 0
        for mrs_file_name in out_text_list:
            temp_flag = 1
            file_object.write('===>' + mrs_file_name + ' :\n')
            for i in range(len(out_text_list[mrs_file_name])):
                file_object.write('\t' + out_text_list[mrs_file_name][i] + '\n')
        if temp_flag == 0:
            file_object.write('test ok')

def mr_test_process():
    mr_utils.conf_xml_parse()

    mr_utils.MR_xml_init()
    #直接写测试需要的16个表单, 对应的函数
    with open(gl.OUT_PATH, 'w') as file_object:
        file_object.write(' '*14 + "<--------------LTE-TEST-RESULT-------------->")
    if gl.TEST_CONF['test51'] == '1':
        try:
            test51_file_integrity()
        except Exception as result:
            write_info ('err in test_51: %s'%(result))
    if gl.TEST_CONF['test52'] == '1':
        try:
            test52_file_integrity()
        except Exception as result:
            write_info ('err in test_52: %s'%(result) )
    if gl.TEST_CONF['test53'] == '1':
        try:
            test53_file_integrity()
        except Exception as result:
            write_info ('err in test_53: %s'% (result) )
    if gl.TEST_CONF['test54'] == '1':
        try:
            test54_file_integrity()
        except Exception as result:
            write_info ('err in test_54: %s'%( result) )
    if gl.TEST_CONF['test55'] == '1':
        try:
            test55_file_integrity()
        except Exception as result:
            write_info ('err in test_55 : %s'%(result))
    if gl.TEST_CONF['test56'] == '1':
        try:
            test56_file_integrity()
        except Exception as result:
            write_info ('err in test_56: %s'%( result))
    if gl.TEST_CONF['test57'] == '1':
        try:
            test57_file_integrity()
        except Exception as result:
            write_info ('err in test_57: %s' %(result) )
    if gl.TEST_CONF['test58'] == '1':
        try:
            test58_file_integrity()
        except Exception as result:
            write_info ('err in test_58: %s' % (result))
    if gl.TEST_CONF['test59'] == '1':
        try:
            test59_file_integrity()
        except Exception as result:
            write_info ('err in test_59: %s' % (result))
    if gl.TEST_CONF['test61'] == '1':
        try:
            test61_file_accuracy()
        except Exception as result:
            write_info ('err in test_61: %s' % (result))
    if gl.TEST_CONF['test62'] == '1':
        try:
            test62_file_accuracy()
        except Exception as result:
            write_info ('err in  test_62: %s'% (result))
    if gl.TEST_CONF['test63'] == '1':
        try:
            test63_file_accuracy()
        except Exception as result:
            write_info ('err in  test_63: %s'% (result))
    if gl.TEST_CONF['test71'] == '1':
        try:
            test71_file_accuracy()
        except Exception as result:
            write_info ('err in test_71: %s'% (result))
    if gl.TEST_CONF['test72'] == '1':
        try:
            test72_file_accuracy()
        except Exception as result:
            write_info ('err in test_72: %s' % (result))
    if gl.TEST_CONF['test73'] == '1':
        try:
            test73_file_accuracy()
        except Exception as result:
            write_info ('err in  test_73: %s' % (result))
    if gl.TEST_CONF['test_add_timestamp'] == '1':
        try:
            test_add_timestamp_number()
        except Exception as result:
            write_info ('err in test add timestamp: %s'% (result))
    try:
        test_add_mro_s_mapping()
    except Exception as result:
        write_info ('err in test add mros mapping : %s'% (result))






